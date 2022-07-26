#!/usr/bin/python
# -*- coding: utf-8 -*-
import os, sys
import openpyxl,xlrd
from os.path import  join,abspath,dirname
import openpyxl
from pymysql.cursors import DictCursor
import flask,json,pymysql,random,re,datetime
import requests
from concurrent.futures import ThreadPoolExecutor
import json
import re
import time
import xmltodict
import random
import datetime
from MongoFunc import DbQuery
from MyExceptions import *
from CommonFunc import CommonFunc
from MysqlFunc import MysqlCommonQuery,MysqlFunc

import time
from Crypto.Cipher import PKCS1_v1_5 as Cipher_pkcs1_v1_5
from Crypto.PublicKey import RSA
import requests
import json
import base64
import random
import datetime
from concurrent.futures import ThreadPoolExecutor
import openpyxl

from Computational_N_N import SQL_report_ods,betting_odds,water_ammount
from Get_excel_path import owner_backer_path

class BetController(object):
    ROBOT_LIBRARY_SCOPE = 'GLOBAL'
    def __init__(self, mysql_info, mongo_info, *args,color='绿色'):
        """
        模拟ctrl给我司推送数据
        """
        BetController.color=color #定义类属性，可以直接使用
        self.session = requests.session()
        self.dbq = DbQuery(mongo_info)
        # self.ctrl_docs = CtrlIoDocs(mysql_info, mongo_info)
        self.cf = CommonFunc()
        self.mysql = MysqlCommonQuery(mysql_info)
        self.my = MysqlFunc(mysql_info)
        self.sr=SQL_report_ods(mysql_info, mongo_info,AB_list,dict)
        self.b0=betting_odds(mysql_info, mongo_info,AB_list,dict)
        self.wt=water_ammount(mysql_info, mongo_info)

    def report_list_Compared01(self,sport_list,sql_list,Compared,excel_report,begin,end,login_account,rold_id):
        print(f"\033[33m登{rold_id}-({login_account})-{excel_report[0]}-{excel_report[1]}：数据对比开始----------------------------------------------------------------------------------------------\033[0m")
        print(f"{self.time(type=excel_report[9])}-查询时间范围：{begin}-{end}")
        if sport_list!= [] and sql_list!= []:
            yyds_list = []
            for key, value in sport_list[0].items():
                yyds_list.append(key)
            for i in range(0, len(sport_list)):
                for j in range(0, len(sql_list)):
                    ppxt_lsit = Compared.split(",")
                    if sport_list[i][ppxt_lsit[0]] == sql_list[j][ppxt_lsit[0]]:
                        if sport_list[i] == sql_list[j]:
                            correct_list.append(sql_list[j])
                            break
                            # print(correct_list[-1])
                            # print(sport_list[i][Compared],"数据对比正确：" + str(sport_list[i]) + "/" + str(sql_list[j]))
                        else:
                            all_yyds=[]
                            for report in range(0, len(sql_list[j])):
                                if sport_list[i][yyds_list[report]] == sql_list[j][yyds_list[report]]:
                                    all_yyds.append(0)
                                elif report == len(sql_list[j]) - 1:
                                    if 1 in all_yyds:
                                        pass
                                    else:
                                        if sql_list[j] in correct_list:
                                            break
                                        else:
                                            correct_list.append(sql_list[j])
                                            break
                                    # print(sport_list[i][Compared], str(yyds_list[report]),"数据对比正确：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]]))
                                else:
                                    if i < len(sportId_list):
                                        yyth = sportId_list[i]
                                    else:
                                        yyth = "参数已用完"
                                    if yyds_list[report]=='options':
                                        num=self.options_report(A=sport_list[i][ppxt_lsit[0]],B='', C=yyds_list[report],D=sport_list[i][yyds_list[report]],E=sql_list[j][yyds_list[report]])
                                        if num==0:
                                            all_yyds.append(0)
                                        else:
                                            all_yyds.append(1)
                                            print("\033[31m" + str(sport_list[i][ppxt_lsit[0]]) + "-" + yyth+str(yyds_list[report]),"数据对比错误(实际结果/预期结果)：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]]) + "\033[0m",type(sport_list[i][yyds_list[report]]),type(sql_list[j][yyds_list[report]]))
                                    else:
                                        all_yyds.append(1)
                                        sport_all_error.append(str(str(sport_list[i][ppxt_lsit[0]]) + "/" +yyth+ str(yyds_list[report]) + "的数据对比错误,请检查SQL查询的字段与接口字段数据是否一致,数据对比：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]])))
                                        print("\033[31m" + str(sport_list[i][ppxt_lsit[0]]) +"-" + yyth, str(yyds_list[report]),"数据对比错误(实际结果/预期结果)：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]]) + "\033[0m",type(sport_list[i][yyds_list[report]]),type(sql_list[j][yyds_list[report]]))

        else:
            if sport_list== []:
                sport_all_error.append("接口数据为空")
            if sql_list== []:
                sport_all_error.append("SQL数据为空")

    def report_list_Compared02(self, sport_list,sql_list, Compared,excel_report,begin,end,login_account,rold_id):
        print(f"\033[33m登{rold_id}-({login_account})-{excel_report[0]}-{excel_report[1]}：数据对比开始----------------------------------------------------------------------------------------------\033[0m")
        print(f"{self.time(type=excel_report[9])}-查询时间范围：{begin}-{end}")
        # print(len(sport_list),len(sql_list))
        # print(sport_list)
        # print(sql_list)
        # print(Compared)
        # print(excel_report)

        if sport_list!=[] and sql_list!=[] :
            yyds_list = []
            if excel_report[0]=='总投注-混合串关-子查询(查询其注单号，包含的比赛)':
                for key, value in sport_list[0][0].items():
                    yyds_list.append(key)
            else:
                for key, value in sport_list[0].items():
                    yyds_list.append(key)
            for i in range(0, len(sport_list)):
                for j in range(0, len(sql_list)):
                    ppxt_lsit =Compared.split(",")
                    if sport_list[i][ppxt_lsit[0]] == sql_list[j][ppxt_lsit[0]]:
                        if sport_list[i][ppxt_lsit[1]] == sql_list[j][ppxt_lsit[1]]:
                            if sport_list[i] == sql_list[j]:
                                correct_list.append(sql_list[j])
                                break
                                # print(str(sport_list[i][ppxt_lsit[0]])+"-"+str(sport_list[i][ppxt_lsit[1]]),"数据对比正确：" + str(sport_list[i]) + "/" + str(sql_list[j]))
                            else:
                                # print(sport_list[i],"\n",sql_list[j])
                                all_yyds=[]
                                for report in range(0, len(sql_list[j])):
                                    if sport_list[i][yyds_list[report]] == sql_list[j][yyds_list[report]]:
                                        all_yyds.append(0)
                                    elif report == len(sql_list[j]) - 1:
                                        if 1 in all_yyds:
                                            pass
                                        else:
                                            if sql_list[j] in correct_list:
                                                break
                                            else:
                                                correct_list.append(sql_list[j])
                                                break
                                        # print(str(sport_list[i][ppxt_lsit[0]])+"-"+str(sport_list[i][ppxt_lsit[1]]), str(yyds_list[report]),"数据对比正确：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]]))
                                    else:
                                        if i<len(sportId_list):
                                            yyth = sportId_list[i]
                                        else:
                                            yyth="参数已用完"
                                        if yyds_list[report] == 'options':
                                            num=self.options_report(A=sport_list[i][ppxt_lsit[0]],B=sport_list[i][ppxt_lsit[1]],C=yyds_list[report],D=sport_list[i][yyds_list[report]],E=sql_list[j][yyds_list[report]])
                                            if num == 0:
                                                all_yyds.append(0)
                                            else:
                                                all_yyds.append(1)
                                                print("\033[31m" + str(sport_list[i][ppxt_lsit[0]]) + "-" + str(sport_list[i][ppxt_lsit[1]] + "-" + yyth),str(yyds_list[report]),"数据对比错误(实际结果/预期结果)：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]]) + "\033[0m",type(sport_list[i][yyds_list[report]]),type(sql_list[j][yyds_list[report]]))

                                        else:
                                            all_yyds.append(1)
                                            sport_all_error.append(str(str(sport_list[i][ppxt_lsit[0]]) + "-" + str(sport_list[i][ppxt_lsit[1]])) + "/" +yyth+ "/" + str(yyds_list[report]) + "的数据对比错误,请检查SQL查询的字段与接口字段数据是否一致,数据对比：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]]))
                                            print("\033[31m" + str(sport_list[i][ppxt_lsit[0]]) + "-" + str(sport_list[i][ppxt_lsit[1]] + "-" + yyth), str(yyds_list[report]),"数据对比错误(实际结果/预期结果)：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]]) + "\033[0m",type(sport_list[i][yyds_list[report]]),type(sql_list[j][yyds_list[report]]))
            if sport_list== []:
                sport_all_error.append("接口数据为空")
            if sql_list== []:
                sport_all_error.append("SQL数据为空")

    def report_list_Compared03(self, sport_list,sql_list, Compared,excel_report,begin,end,login_account,rold_id):
        print(f"\033[33m登{rold_id}-({login_account})-{excel_report[0]}-{excel_report[1]}：数据对比开始----------------------------------------------------------------------------------------------\033[0m")
        print(f"{self.time(type=excel_report[9])}-查询时间范围：{begin}-{end}")
        if sport_list!= [] and sql_list!= []:
            yyds_list = []
            for key, value in sport_list[0].items():
                yyds_list.append(key)
            for i in range(0, len(sql_list)):
                for j in range(0, len(sql_list)):
                    str_sum = Compared
                    ppxt_lsit = str_sum.split(",")
                    # print(ppxt_lsit)
                    if sport_list[i][ppxt_lsit[0]] == sql_list[j][ppxt_lsit[0]]:
                        if sport_list[i][ppxt_lsit[1]] == sql_list[j][ppxt_lsit[1]]:
                            if sport_list[i][ppxt_lsit[2]] == sql_list[j][ppxt_lsit[2]]:
                                if sport_list[i] == sql_list[j]:
                                    correct_list.append(sql_list[j])
                                    break
                                    # print(str(sport_list[i][ppxt_lsit[0]])+"-"+str(sport_list[i][ppxt_lsit[1]])+"-"+str(sport_list[i][ppxt_lsit[2]]),"数据对比正确：" + str(sport_list[i]) + "/" + str(sql_list[j]))
                                else:
                                    all_yyds = []
                                    for report in range(0, len(sql_list[j])):
                                        if sport_list[i][yyds_list[report]] == sql_list[j][yyds_list[report]]:
                                            all_yyds.append(0)
                                        elif report == len(sql_list[j]) - 1:
                                            if 1 in all_yyds:
                                                pass
                                            else:
                                                if sql_list[j] in correct_list:
                                                    break
                                                else:
                                                    correct_list.append(sql_list[j])
                                                    break
                                            # print(str(sport_list[i][ppxt_lsit[0]])+"-"+str(sport_list[i][ppxt_lsit[1]])+"-"+str(sport_list[i][ppxt_lsit[2]]), str(yyds_list[report]),"数据对比正确：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]]))
                                        else:
                                            if i < len(sportId_list):
                                                yyth = sportId_list[i]
                                            else:
                                                yyth = "参数已用完"
                                            if yyds_list[report] == 'options':
                                                num = self.options_report(A=sport_list[i][ppxt_lsit[0]],B=sport_list[i][ppxt_lsit[1]],C=yyds_list[report],D=sport_list[i][yyds_list[report]],E=sql_list[j][yyds_list[report]])
                                                if num == 0:
                                                    all_yyds.append(0)
                                                else:
                                                    all_yyds.append(1)
                                                    print("\033[31m" + str(sport_list[i][ppxt_lsit[0]]) + "-" + str(sport_list[i][ppxt_lsit[1]]) + "-" + str(sport_list[i][ppxt_lsit[2]])+ "-"+ yyth),str(yyds_list[report]), "数据对比错误(实际结果/预期结果)：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]]) + "\033[0m",type(sport_list[i][yyds_list[report]]),type(sql_list[j][yyds_list[report]])
                                            else:
                                                all_yyds.append(1)
                                                sport_all_error.append(str(str(sport_list[i][ppxt_lsit[0]]) + "-" + str(sport_list[i][ppxt_lsit[1]])) + "-"  + str(sport_list[i][ppxt_lsit[2]])+ "/" + yyth + "/" + str(yyds_list[report]) + "的数据对比错误,请检查SQL查询的字段与接口字段数据是否一致,数据对比：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]]))
                                                print("\033[31m" + str(sport_list[i][ppxt_lsit[0]]) + "-" + str(sport_list[i][ppxt_lsit[1]] + "-"  + str(sport_list[i][ppxt_lsit[2]])+ "-"+ yyth),str(yyds_list[report]), "数据对比错误(实际结果/预期结果)：" + str(sport_list[i][yyds_list[report]]) + "/" + str(sql_list[j][yyds_list[report]]) + "\033[0m",type(sport_list[i][yyds_list[report]]),type(sql_list[j][yyds_list[report]]))

        else:
            if sport_list == []:
                sport_all_error.append("接口数据为空")
            if sql_list== []:
                sport_all_error.append("SQL数据为空")

    def options_report(self,A, B, C,D,E):
        options_yyds=0
        aa_list = []
        # print(D[0])
        # print(E[0])
        tyyt=0
        for key, value in D[0].items():
            aa_list.append(key)
        if D==E:
            print("options字典数据对比正确")
        else:
            if len(D)==len(E):
                for ykk in range(0, len(D)):
                    for ytt in range(0, len(E)):
                        if D[ykk][aa_list[0]]==E[ytt][aa_list[0]]:
                            if D[ykk]==E[ytt]:
                                pass
                                # print(f"{A}-{B},{C}options数据对比正确：{D[ykk]}/{E[ytt]}")
                            else:
                                 for i in aa_list:
                                     if D[ykk][i] == E[ytt][i]:
                                         # print(f"{A}-{B},{C}options数据对比正确：{D[ykk][i]}/{E[ytt][i]}")
                                         options_yyds=0
                                     else:
                                         if aa_list.index(i)==0:
                                             print(f"\033[32m{A}-{B},{C}：options字典数据对比开始---------------------------------------------------------------------------\033[0m")
                                         print(f"\033[31m{A}-{B},{C}options数据对比错误：{D[ykk][i]}/{E[ytt][i]}\033[0m{type(D[ykk][i])}/{type(E[ytt][i])}-")
                                         if aa_list.index(i)==aa_list[-1]:
                                             print(f"\033[32m{A}-{B},{C}：options字典数据对比结束----------------------------------------------------------------------------------------------\033[0m")
                                         options_yyds=1

                        else:
                            # if ytt==len(E)-1:
                            #     if tyyt==0:
                            #         print(f"\033[31m{A}-{B},{C}options数据对比错误：{D}/{E}\033[0m")
                            #         tyyt=tyyt+1
                            continue
            else:
                print(f"\033[32m{A}-{B},{C}：options字典数据对比开始---------------------------------------------------------------------------\033[0m")
                print(f"\033[31m{A}-{B},{C}数据字典数量不一致错误:{len(D)}/{len(E)}")
                print(f"\033[32m{A}-{B},{C}：options字典数据对比结束----------------------------------------------------------------------------------------------\033[0m")
        return options_yyds


    def if_error(self, A,B,correct_list):
        key01=[]
        key02 = []
        error=''
        number_num=''
        if len(A)>len(B):
            number_num=0
        elif len(A)==len(B):
            number_num = random.randint(0, len(B) - 1)
        elif len(A)!=len(B):
            number_num = random.randint(0, len(A) - 1)
        if len(A[number_num])==len(B[number_num]):
            for key,value in A[number_num].items():
                key01.append(key)
            if 'options' in key01:
                if len((A[number_num]['options'][0]).keys())==len((B[number_num]['options'][0]).keys()):
                    yytr=self.options_report(A=A[number_num][key01[0]], B=A[number_num][key01[-1]], C='options',D=A[number_num]['options'],E=B[number_num]['options'])
                    if yytr==0:
                        if A==B:
                            print(f"\33[32m接口数据数据预校验通过\33[0m")
                            correct_list=B
                        else:
                            print("\33[31m数据预校验不通过\33[0m")
                            error = '错误'
                    else:
                        print(f"\33[31m字典value的值不相等\33[0m")
                        error = '错误'
                else:
                    print(f"A与B，key值不一致{len((A[number_num]['options'][0]).keys())}/{len((B[number_num]['options'][0]).keys())}")
                    for key, value in B[0]['options'].items():
                        key02.append(key)
                    for key_value in key02:
                        if key_value in key01:
                            del key01[key01.index(key_value)]
                        else:
                            pass
                    print(f"\33[31mA与B，options字典数量不相等,他们相差{key01}\33[0m")
            else:
                if A==B:
                    print(f"\33[32m数据预校验通过\33[0m")
                    correct_list=B
                else:
                    print(f"\33[31m数据预校验不通过\33[0m")
                    error='错误'
        else:
            print(f"\33[31m接口数据与SQL数据字段比对错误,A:{len(A[number_num])}/B:{len(B[number_num])}\33[0m")
            error = '错误'
        return error,correct_list




    def null_data_processing(self,sort_num,mix_number,excel_report):
        print(f"\33[33m{login_account}-正在检验{excel_report[0]}-{excel_report[1]}数据是否返回为空\33[0m")
        print(sort_num,len(sort_num))
        report_null=''
        if sort_num==() or sort_num==None or sort_num[0]==() or sort_num[0]==None or sort_num[0][0]==() or sort_num[0][0]==None:
            report_null = []
        else:
            report_null=sort_num
        return report_null

    def null_report(self,sport_report_list,sql_report_list,excel_report):
        global Prepare_enough_list,Prepare_enough_list,sport_error
        Prepare_enough_list=[]
        sport_error=[]
        # if sport_report_list==[]:
        #     sql_report_list=[]
        # else:
        #     pass

        if sport_report_list==sql_report_list:
            print(f"\033[32m代理{sport_report_list}-{sql_report_list}比对SQL数据一致,都是空列表\033[0m")
            Prepare_enough_list.append(f"代理{sport_report_list}-{sql_report_list}与SQL数据一致,都是空列表")
        else:
            print(f"\033[31m代理{excel_report[0]}-{excel_report[1]}，比对SQL数据错误，需要比对\033[0m")
            print(f"代理查询{sport_report_list}-SQL查询{sql_report_list}，比对SQL数据错误，需要比对")
            print("\n","\n")
            Prepare_enough_list.append(f"代理查询{sport_report_list}-SQL查询{sql_report_list}，比对SQL数据错误，需要比对")
            sport_error.append(sport_report_list)
            sport_error.append(sql_report_list)

    #判断读取的SQL内容，是否要写入数据
    def if_f(self,BBQ,all,num,all_order_no,begin,end,login_account,rold_id):
        #判断SQL是否写入
        if BBQ[0] == "f":
            sql=eval(BBQ)
        else:
            sql=BBQ
        return sql

    #对已经组装好的比赛数据，对于相同订单号组装在一起，不同订单号分开组装
    def order_no_new(self,yyds_list):
        global orderNo_list
        orderNo_list = []
        for i in yyds_list:
            orderNo_tuple = []
            #p判断订单号是否在列表中，在的，直接在原列表中添加元素，不在直接添加新的元素
            if i['orderNo'] in orderNo_list:
                index_number = orderNo_list.index(i['orderNo'])
                options_list[index_number].append(i)
            else:
                orderNo_list.append(i['orderNo'])
                orderNo_tuple.append(i)
                options_list.append(orderNo_tuple)
        return options_list

    #对多个数据的订单号，单独进行组装
    def order_no_orderno(self,yyds_list,excel_report):

        # new_orderNo_list = []
        orderNo_list = []
        for i in yyds_list:
            orderNo_tuple = []
            index_i=0
            #p判断订单号是否在列表中，在的，直接在原列表中添加元素，不在直接添加新的元素
            if i[index_i] in orderNo_list:
                index_number = orderNo_list.index(i[index_i])
                if excel_report[0]=="总投注-混合串关-子查询(查询其注单号，包含的比赛)":
                    i=i[1:]
                options_list[index_number].append(i)
            else:
                orderNo_list.append(i[index_i])
                if excel_report[0]=="总投注-混合串关-子查询(查询其注单号，包含的比赛)":
                    i=i[1:]
                orderNo_tuple.append(i)
                options_list.append(orderNo_tuple)
        return options_list



    def sport_report_sql(self,begin,end,excel_report,sport_report_dict,sport_report_list,login_account,rold_id):
        global options_list,sportId_sql_list,sport_all_error
        print(f"\33[35m登{rold_id}-({login_account})-正在执行SQL{excel_report[0]}-{excel_report[1]}\33[0m")

        num_list = []
        sort_num_list = []
        mix_number=[]
        options_list = []
        sportId_sql_list = []
        sport_num_sql_list=[]
        sql_yyds =[]
        sort_num_list=[]
        sport_all_error = []
        marketID_sql_list = []
        marketID_sql_dict={}


        sum_yyds = excel_report[11]
        yyds = sum_yyds.split("@")
        report_null=''
        all=''
        num=''
        sql_count=0
        all_order_no=''
        rold_id=rold_id
        login_account=login_account
        if yyds[0]=='':
            del yyds[0]
        if excel_report[3] == 1:
            if yyds[0]=='':
                if excel_report[0] in ('首页-余额详细资料-主查询', '首页-统计-主查询'):
                    if excel_report[0] == '首页-余额详细资料-主查询':
                        sort_num = self.wt.Home_Balance_Details(agent_id=login_account)
                        print(sort_num)
                        mix_number.append("1-1")
                        report_null = self.null_data_processing(sort_num=sort_num, mix_number=mix_number,excel_report=excel_report)
                        if report_null == []:
                            self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                        else:
                            pass
            elif len(yyds) == 1:
                sql=self.if_f(BBQ=yyds[0],all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
                sort_num = self.my.query_data(sql, db_name='bfty_credit')
                mix_number.append("1-1")
                print(sql)
                report_null=self.null_data_processing(sort_num=sort_num,mix_number=mix_number,excel_report=excel_report)
                if report_null==[]:
                    self.null_report(sport_report_list=sport_report_list,sql_report_list=report_null,excel_report=excel_report)
                else:
                    pass
            else:
                if excel_report[0] in ('报表-已取消注单-主查询'):
                    sql01 = self.if_f(BBQ=yyds[0], all=all, num=num, all_order_no=all_order_no, begin=begin, end=end,rold_id=rold_id,login_account=login_account)
                    sql02 = self.if_f(BBQ=yyds[1], all=all, num=num, all_order_no=all_order_no, begin=begin, end=end,rold_id=rold_id,login_account=login_account)
                    print(sql01)
                    print(sql02)
                    sort_num01 = self.my.query_data(sql01, db_name='bfty_credit')
                    sort_num02 = self.my.query_data(sql02, db_name='bfty_credit')
                    # 单独获取options列表的数据，然后组装回去
                    aa_list = []
                    jkk_list = []
                    #判断接口数据为空时，不在循环
                    if sport_report_list == []:
                        report_null = self.null_data_processing(sort_num=sort_num01, mix_number=mix_number,excel_report=excel_report)
                        self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                    else:
                        report_null = self.null_data_processing(sort_num=sort_num01, mix_number=mix_number,excel_report=excel_report)
                        if report_null == []:
                            self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                        else:
                            for key, value in sport_report_list[0]['options'][0].items():
                                aa_list.append(key)
                            for jkk in sort_num02:
                                jkk_dict = {}
                                for jkk_2 in range(0, len(jkk)):
                                    if str(type(jkk[jkk_2])) in type_list:
                                        if jkk[jkk_2] == None:
                                            new_jkk = None
                                        else:
                                            new_jkk = str(jkk[jkk_2])
                                    else:
                                        new_jkk = float(jkk[jkk_2])
                                    jkk_dict[aa_list[jkk_2]] = new_jkk
                                jkk_list.append(jkk_dict)
                            self.order_no_new(yyds_list=jkk_list)
                        sort_num = sort_num01
                        mix_number.append("1-1")
                else:
                    if excel_report[0]=='用户管理-下级结账-主查询':
                        sql01 = self.if_f(BBQ=yyds[0], all=all, num=num, all_order_no=all_order_no, begin=begin,end=end)
                        sort_num01 = self.my.query_data(sql01, db_name='bfty_credit')
                        num=int(sort_num01[0][0])+1
                        sql02 = self.if_f(BBQ=yyds[1], all=all, num=num, all_order_no=all_order_no, begin=begin,end=end)
                        sort_num02 = self.my.query_data(sql02, db_name='bfty_credit')
                        print(sql01, sql02)
                        sort_num_list.append(sort_num02)
                        num_list.append(sort_num_list)
                        sort_num = num_list
                        mix_number.append("2-2")
                    else:
                        sql01 =self.if_f(BBQ=yyds[0],all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
                        sql02 =self.if_f(BBQ=yyds[1],all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
                        sort_num01 = self.my.query_data(sql01, db_name='bfty_credit')
                        sort_num02 = self.my.query_data(sql02, db_name='bfty_credit')
                        report_null = self.null_data_processing(sort_num=sort_num02, mix_number=mix_number,excel_report=excel_report)
                        if report_null == []:
                            pass
                        else:
                            sort_num_list.append(sort_num02)
                        sort_num_list.append(sort_num01)
                        print(sql01,sql02)

                        num_list.append(sort_num_list)
                        sort_num = num_list
                        mix_number.append("2-2")
                    report_null = self.null_data_processing(sort_num=sort_num,mix_number=mix_number,excel_report=excel_report)
                    if report_null == []:
                        self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                    else:
                        pass
        elif excel_report[3] == 2:
            if excel_report[0] in ('总投注-混合串关-子查询(查询其注单号，包含的比赛)',"未完成交易-登0-登3-会员-查看订单详情-子查询","盈亏详情-登0-登3-会员-查看订单详情-子查询","报表-球类报表-订单查询（根据其盘口查询订单）","报表-联赛报表-订单查询（根据其盘口查询订单）",'报表-赛事盈亏-订单查询（根据其盘口查询订单）','报表-混合串关-订单查询（根据其会员查询订单）','盈亏详情-登1-登3-会员-查看订单详情-子查询','盈亏详情-登2-登3-会员-查看订单详情-子查询','盈亏详情-登3-会员-查看订单详情-子查询'):
                if excel_report[0]=='总投注-混合串关-子查询(查询其注单号，包含的比赛)':
                    sql = self.if_f(BBQ=yyds[0], all=all, num=num, all_order_no=all_order_no, begin=begin, end=end,rold_id=rold_id,login_account=login_account)
                    sort_num = self.my.query_data(sql, db_name='bfty_credit')
                    num_list = self.order_no_orderno(yyds_list=sort_num, excel_report=excel_report)
                    mix_number.append("2-1")
                    print(sql)
                    sort_num = num_list
                    report_null = self.null_data_processing(sort_num=sort_num, mix_number=mix_number,excel_report=excel_report)
                    if report_null == []:
                        self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                    else:
                        pass
                elif excel_report[0] in ('报表-球类报表-订单查询（根据其盘口查询订单）','报表-联赛报表-订单查询（根据其盘口查询订单）','报表-赛事盈亏-订单查询（根据其盘口查询订单）','报表-混合串关-订单查询（根据其会员查询订单）'):
                    sql01 = self.if_f(BBQ=yyds[0], all=all, num=num, all_order_no=all_order_no, begin=begin, end=end,rold_id=rold_id,login_account=login_account)
                    sql02 = self.if_f(BBQ=yyds[1], all=all, num=num, all_order_no=all_order_no, begin=begin, end=end,rold_id=rold_id,login_account=login_account)
                    print(sql01)
                    print(sql02)
                    sort_num01 = self.my.query_data(sql01, db_name='bfty_credit')
                    sort_num02 = self.my.query_data(sql02, db_name='bfty_credit')

                    # 判断接口数据为空时，不在循环
                    if sport_report_list == []:
                        report_null = self.null_data_processing(sort_num=sort_num01, mix_number=mix_number,excel_report=excel_report)
                        self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                    else:
                        # 单独获取options列表的数据，然后组装回去
                        aa_list = []
                        jkk_list = []
                        for key, value in sport_report_list[0]['options'][0].items():
                            aa_list.append(key)
                        for jkk in sort_num02:
                            jkk_dict = {}
                            for jkk_2 in range(0, len(jkk)):
                                if str(type(jkk[jkk_2])) in type_list:
                                    if jkk[jkk_2] == None:
                                        new_jkk = None
                                    else:
                                        new_jkk = str(jkk[jkk_2])
                                else:
                                    new_jkk = float(jkk[jkk_2])
                                jkk_dict[aa_list[jkk_2]] = new_jkk
                            jkk_list.append(jkk_dict)
                        self.order_no_new(yyds_list=jkk_list)
                        mix_number.append("1-1")
                        sort_num = sort_num01
                elif excel_report[0] in ('未完成交易-登0-登3-会员-查看订单详情-子查询',"盈亏详情-登0-登3-会员-查看订单详情-子查询",'盈亏详情-登1-登3-会员-查看订单详情-子查询','盈亏详情-登2-登3-会员-查看订单详情-子查询','盈亏详情-登3-会员-查看订单详情-子查询'):
                    sql01 = self.if_f(BBQ=yyds[0], all=all, num=num, all_order_no=all_order_no, begin=begin, end=end,rold_id=rold_id,login_account=login_account)
                    sql02 = self.if_f(BBQ=yyds[1], all=all, num=num, all_order_no=all_order_no, begin=begin, end=end,rold_id=rold_id,login_account=login_account)
                    sql03 = self.if_f(BBQ=yyds[2], all=all, num=num, all_order_no=all_order_no, begin=begin, end=end,rold_id=rold_id,login_account=login_account)
                    print(sql01)
                    print(sql02)
                    print(sql03)
                    sort_num01 = self.my.query_data(sql01, db_name='bfty_credit')
                    sort_num02 = self.my.query_data(sql02, db_name='bfty_credit')
                    sort_num03 = self.my.query_data(sql03, db_name='bfty_credit')

                    # 判断接口数据为空时，不在循环
                    if sport_report_list == []:
                        report_null = self.null_data_processing(sort_num=sort_num01, mix_number=mix_number,excel_report=excel_report)
                        self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                    else:
                        #把sort_num01的总金额数据，组装进sort_num02的里
                        user_name01=[]
                        user_money=[]
                        #先把总金额数据，登入账号和金额取出来，分别放入user_name01，user_money
                        for j in sort_num01:
                            user_name01.append(j[0])
                            if excel_report[0]=='未完成交易-登0-登3-会员-查看订单详情-子查询':
                                user_money.append(float(j[1]))
                            else:
                                user_money.append((j[1:]))
                        #写入信息中
                        for i in range(0,len(sort_num02)):
                            ppt_list=[]
                            if sort_num02[i][0] in user_name01:
                                if excel_report[0] == '未完成交易-登0-登3-会员-查看订单详情-子查询':
                                    new_report=sort_num02[i]+((user_money[user_name01.index(sort_num02[i][0])]),)
                                else:
                                    new_report=sort_num02[i]
                                    for jj in (user_money[user_name01.index(sort_num02[i][0])]):
                                        new_report = new_report + ((jj),)
                                ppt_list.append(new_report)
                                num_list.append(ppt_list)
                        # 单独获取options列表的数据，然后组装回去
                        aa_list = []
                        jkk_list = []
                        for key, value in sport_report_list[0]['options'][0].items():
                            aa_list.append(key)
                        for jkk in sort_num03:
                            jkk_dict = {}
                            for jkk_2 in range(0, len(jkk)):
                                if str(type(jkk[jkk_2])) in type_list:
                                    if jkk[jkk_2] == None:
                                        new_jkk = None
                                    else:
                                        new_jkk = str(jkk[jkk_2])
                                else:
                                    new_jkk = float(jkk[jkk_2])
                                jkk_dict[aa_list[jkk_2]] = new_jkk
                            jkk_list.append(jkk_dict)
                        self.order_no_new(yyds_list=jkk_list)
                        mix_number.append("2-1")
                        sort_num = num_list
            else:
                sql011=self.if_f(BBQ=yyds[0],all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
                sort_num = self.my.query_data(sql011, db_name='bfty_credit')
                print(sql011)
                del yyds[0]
                if sport_report_list==[]:
                    report_null = self.null_data_processing(sort_num=sort_num, mix_number=mix_number,excel_report=excel_report)
                    self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                else:
                    report_null = self.null_data_processing(sort_num=sort_num, mix_number=mix_number, excel_report=excel_report)
                    if report_null == []:
                        self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                    for sportId in sort_num:
                        sportId_sql_list.append(sportId[0])
                    for all in sportId_sql_list:
                        sort_num_list = []
                        if len(yyds) == 1:
                            sql=self.if_f(BBQ=yyds[0],all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
                            sort_num = self.my.query_data(sql, db_name='bfty_credit')
                            num_list.append(sort_num)
                            if num_list[0] ==():
                                del num_list[0]
                            sort_num = num_list
                            mix_number.append("2-1")
                        elif len(yyds) == 2:
                            sql01 = self.if_f(BBQ=yyds[0],all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
                            sql02 = self.if_f(BBQ=yyds[1],all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
                            sort_num01 = self.my.query_data(sql01, db_name='bfty_credit')
                            sort_num02 = self.my.query_data(sql02, db_name='bfty_credit')
                            sort_num_list.append(sort_num01)
                            sort_num_list.append(sort_num02)
                            num_list.append(sort_num_list)
                            sort_num = num_list
                            mix_number.append("2-2")
                            print(sql01,sql02)
                            sql_count = sql_count + 1
                    report_null = self.null_data_processing(sort_num=sort_num, mix_number=mix_number,excel_report=excel_report)
                    if report_null == []:
                        self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                    else:
                        pass

        else:
            sql011=self.if_f(BBQ=yyds[0],all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
            print(sql011)
            sort_num = self.my.query_data(sql011, db_name='bfty_credit')
            if sport_report_list == []:
                report_null = self.null_data_processing(sort_num=sort_num, mix_number=mix_number,excel_report=excel_report)
                self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
            else:
                report_null = self.null_data_processing(sort_num=sort_num, mix_number=mix_number,excel_report=excel_report)
                if report_null == []:
                    self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                else:
                    for all in sort_num[0]:
                        sql012 = self.if_f(BBQ=yyds[1],all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
                        sort_num = self.my.query_data(sql012, db_name='bfty_credit')
                        sql_yyds.append(sort_num)
                    del yyds[0:2]
                    for sportId in sql_yyds:
                        sportId_sql_list.append(sportId[0][0])
                        sport_num_sql_list.append(sportId[0][1])
                    for all in sportId_sql_list:
                        sort_num_list = []
                        if len(yyds) == 1:
                            if excel_report[0] in interface_list:
                                sort_num = ''
                                num=int(sport_num_sql_list[sportId_sql_list.index(all)])
                                if num==0:
                                    continue
                                else:
                                    sql = self.if_f(BBQ=eval(yyds[0]),all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
                                    sort_num = self.my.query_data(sql, db_name='bfty_credit')
                                    num_list.append(sort_num)
                                    mix_number.append("2-1")
                                    sql_count = sql_count + 1
                        else:
                            sql01 = self.if_f(BBQ=yyds[0],all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
                            sql02 = self.if_f(BBQ=yyds[1],all=all,num=num,all_order_no=all_order_no,begin=begin,end=end,rold_id=rold_id,login_account=login_account)
                            sort_num01 = self.my.query_data(sql01, db_name='bfty_credit')
                            sort_num02 = self.my.query_data(sql02, db_name='bfty_credit')
                            sort_num_list.append(sort_num01)
                            sort_num_list.append(sort_num02)
                            num_list.append(sort_num_list)
                            mix_number.append("2-2")
                            sql_count = sql_count + 1
                            # print(len(sport_report_list), sql_count)
                    sort_num = num_list
                    report_null = self.null_data_processing(sort_num=sort_num, mix_number=mix_number, excel_report=excel_report)
                    if report_null == []:
                        self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
                    else:
                        pass




        sport_sql_list = sport_report_dict
        sql_sport_list = []
        # print(sort_num[0:2])
        # print(options_list)
        # print(f"SQL数据：{sort_num}\n接口数据：{sport_report_list}\n-----------------------------------------------------------------------------")
        print("-------------------------------------------------------------------------------------------------------")
        if sport_report_list==[] or report_null == []:
            self.null_report(sport_report_list=sport_report_list, sql_report_list=report_null,excel_report=excel_report)
        else:
            self.Compared_report(sort_num=sort_num,sport_list=sport_report_list,sql_name_list=sport_sql_list,sql_list=sql_sport_list,Compared=excel_report[12],report_name=excel_report[0]+"  "+excel_report[1],excel_report=excel_report,mix=mix_number,begin=begin,end=end,login_account=login_account,rold_id=rold_id)

    def Compared_report(self, sort_num, sport_list, sql_name_list, sql_list, Compared, report_name, excel_report,mix,begin,end,login_account,rold_id):
        global sport_error,yyds01,correct_list,Prepare_enough_list
        #调试
        # print(sort_num[0])
        # exit()
        correct_list=[]
        Prepare_enough_list=[]
        sport_error = []
        str_num=Compared
        yyds01=str_num.split(",")
        if mix==[]:
            pass
        else:
            print(mix[0])
        if mix[0]==str('1-1'):
            print(sort_num[0])
            # print(sort_num)
            print(len(sport_list),len(sort_num), len(options_list))
            for i in range(0, len(sort_num)):
                sql_dict = {}
                for j in range(0, len(sort_num[i])):
                    if str(type(sort_num[i][j])) in type_list:
                        if sort_num[i][j]=="odds":
                            if excel_report[0]  in ('报表-球类报表-订单查询（根据其盘口查询订单）','报表-联赛报表-订单查询（根据其盘口查询订单）','报表-赛事盈亏-订单查询（根据其盘口查询订单）','报表-混合串关-订单查询（根据其会员查询订单）','报表-已取消注单-主查询','总投注-混合串关-主查询'):
                                if excel_report[0]=='报表-混合串关-订单查询（根据其会员查询订单）':
                                    yy_num = float(self.sr.credit_odds(order_no=sort_num[i][-6], bet_type="", AB_list=AB_list,dict=dict))
                                elif excel_report[0]=='报表-已取消注单-主查询':
                                    yy_num = float(self.sr.credit_odds(order_no=sort_num[i][-4], bet_type="", AB_list=AB_list,dict=dict))
                                elif excel_report[0]=='总投注-混合串关-主查询':
                                    yy_num = self.sr.credit_odds(order_no=sort_num[i][-3], bet_type="", AB_list=AB_list,dict=dict)
                                else:
                                    yy_num =float(self.sr.credit_odds(order_no=sort_num[i][35], bet_type="", AB_list=AB_list,dict=dict))
                        elif sort_num[i][j]=="odds_type":
                            if excel_report[0] in ('报表-球类报表-订单查询（根据其盘口查询订单）','报表-联赛报表-订单查询（根据其盘口查询订单）','报表-赛事盈亏-订单查询（根据其盘口查询订单）','报表-混合串关-订单查询（根据其会员查询订单）','报表-已取消注单-主查询','总投注-混合串关-主查询'):
                                if excel_report[0] == '报表-混合串关-订单查询（根据其会员查询订单）':
                                    yy_num = str(self.sr.total_odds(order_no=sort_num[i][-6], number=' '))
                                elif excel_report[0] == '报表-已取消注单-主查询':
                                    yy_num = str(self.sr.total_odds(order_no=sort_num[i][-4], number=' '))
                                elif excel_report[0] == '总投注-混合串关-主查询':
                                    yy_num = self.sr.total_odds(order_no=sort_num[i][-3],number='str')
                                else:
                                    yy_num = str(self.sr.total_odds(order_no=sort_num[i][35], number=' '))
                        elif sort_num[i][j]=="options":
                            yy_num = options_list[i]
                        elif sql_name_list[j] in int_list:
                            if sql_name_list[j]=="levelId":
                                yy_num = int(sort_num[i][j])+1
                            else:
                                if excel_report[0] in ('报表-球类报表-订单查询（根据其盘口查询订单）','报表-联赛报表-订单查询（根据其盘口查询订单）','报表-赛事盈亏-订单查询（根据其盘口查询订单）','报表-混合串关-订单查询（根据其会员查询订单）','报表-已取消注单-主查询','总投注-混合串关-主查询'):
                                    yy_num=(sort_num[i][j])
                                else:
                                    yy_num = int(sort_num[i][j])
                        elif sort_num[i][j] == "公司总计":
                            if excel_report[0] in options_report_list:
                                yy_num = self.wt.Company_winlose(agent_id='',member_id=sort_num[i][1],sportId='',marketId='',tournamentId='',matchId='',login_account=login_account,begin=begin,end=end,Duplex='',excel_num=excel_report[9])
                            elif  excel_report[0] in one_report_list:
                                if excel_report[0] in("报表-球类报表-主查询","报表-球类报表-主查询-投注时间","报表-球类报表-主查询-赛事时间","报表-球类报表-主查询-结算时间",):
                                    yy_num = self.wt.Company_winlose(agent_id='',member_id='',sportId=sort_num[i][-2],marketId='',tournamentId='',matchId='',login_account=login_account,begin=begin,end=end,Duplex='',excel_num=excel_report[9])
                                elif excel_report[0] in ("报表-混合串关-主查询","报表-混合串关-主查询-投注时间","报表-混合串关-主查询-赛事时间","报表-混合串关-主查询-结算时间",):
                                    yy_num = self.wt.Company_winlose(agent_id='', member_id=sort_num[i][-1], sportId='',marketId='', tournamentId='', matchId='',login_account=login_account, begin=begin, end=end,Duplex=1,excel_num=excel_report[9])
                        elif sort_num[i][j] == "总佣金":
                            if excel_report[0] in options_report_list:
                                yy_num = self.wt.total_commission(agent_id='', member_id=sort_num[i][1],sportId='',marketId='',tournamentId='',matchId='',login_account=login_account,begin=begin,end=end,Duplex='',ZD='',excel_num=excel_report[9])
                            elif  excel_report[0] in one_report_list:
                                if excel_report[0] in("报表-球类报表-主查询","报表-球类报表-主查询-投注时间","报表-球类报表-主查询-赛事时间","报表-球类报表-主查询-结算时间",):
                                    yy_num = self.wt.total_commission(agent_id='',member_id='',sportId=sort_num[i][-2],marketId='',tournamentId='',matchId='',login_account=login_account,begin=begin,end=end,Duplex='',ZD='',excel_num=excel_report[9])
                                elif excel_report[0] in ("报表-混合串关-主查询","报表-混合串关-主查询-投注时间","报表-混合串关-主查询-赛事时间","报表-混合串关-主查询-结算时间",):
                                    yy_num = self.wt.total_commission(agent_id='', member_id=sort_num[i][-5],sportId='', marketId='', tournamentId='',matchId='', login_account=login_account,begin=begin, end=end,Duplex=1,ZD='',excel_num=excel_report[9])
                        else:
                            yy_num=str(sort_num[i][j])
                    else:
                        yy_num = float(sort_num[i][j])
                    sql_dict[sql_name_list[j]] =yy_num
                sql_list.append(sql_dict)
            if len(sport_list)>1:
                print("sport_list组装数据对比：", sport_list[1])
                print("sql_list组装数据对比：", sql_list[1])
            else:
                print("sport_list组装数据对比：", sport_list)
                print("sql_list组装数据对比：", sql_list)
        if mix[0]==str('2-1'):
            # print(sort_num)
            print(sort_num[0])
            print(sort_num[0][0])
            print(len(sport_list), len(sort_num), len(options_list))
            count = 0
            for i in range(0, len(sort_num)):
                for j in range(0,len(sort_num[i])):
                    sql_dict = {}
                    print(sort_num[i][j])
                    for g in range(0,len(sort_num[i][j])):
                        # exit(f"{len(sort_num)},{len(options_list)}")
                        if str(type(sort_num[i][j][g]))in type_list:
                            if sql_name_list[g] in int_list:
                                if excel_report[0] in ('盈亏详情-登0-登3-会员-查看订单详情-子查询','报表-球类报表-订单查询（根据其盘口查询订单）','盈亏详情-登1-登3-会员-查看订单详情-子查询','盈亏详情-登2-登3-会员-查看订单详情-子查询','盈亏详情-登3-会员-查看订单详情-子查询'):
                                    yy_num = str(sort_num[i][j][g])
                                else:
                                    if str(type(sort_num[i][j][g]))=="<class 'datetime.datetime'>":
                                        yy_num = str(sort_num[i][j][g])
                                    else:
                                        yy_num = int(sort_num[i][j][g])
                            elif sort_num[i][j][g]=="odds":
                                if excel_report[0]=='未完成交易-登0-登3-会员-查看订单详情-子查询':
                                        yy_num =float(self.sr.credit_odds(order_no=sort_num[i][j][-3], bet_type="",AB_list=AB_list, dict=dict))
                                elif excel_report[0] in ('盈亏详情-登0-登3-会员-查看订单详情-子查询','盈亏详情-登1-登3-会员-查看订单详情-子查询','盈亏详情-登2-登3-会员-查看订单详情-子查询','盈亏详情-登3-会员-查看订单详情-子查询'):
                                        yy_num =float(self.sr.credit_odds(order_no=sort_num[i][j][-23], bet_type="",AB_list=AB_list, dict=dict))
                            elif sort_num[i][j][g] == "options":
                                if excel_report[0]=="未完成交易-登0-登3-会员-查看订单详情-子查询":
                                    if sort_num[i][j][-3] in orderNo_list:
                                        yy_num = options_list[orderNo_list.index(sort_num[i][j][-3])]
                                else:
                                    yy_num = options_list[i]
                            elif sort_num[i][j][g] == "odds_type":
                                if excel_report[0] == '未完成交易-登0-登3-会员-查看订单详情-子查询':
                                    yy_num = self.sr.total_odds(order_no=sort_num[i][j][-3], number=' ')
                                elif excel_report[0] in ('盈亏详情-登0-登3-会员-查看订单详情-子查询','盈亏详情-登1-登3-会员-查看订单详情-子查询','盈亏详情-登2-登3-会员-查看订单详情-子查询','盈亏详情-登3-会员-查看订单详情-子查询'):
                                    yy_num = self.sr.total_odds(order_no=sort_num[i][j][-23], number=' ')
                            elif sort_num[i][j][g] == "公司总计":
                                if excel_report[0] in interface_list:
                                    yy_num = self.wt.Company_winlose(agent_id=sort_num[i][j][1], member_id='', sportId='',marketId='', tournamentId='', matchId='',login_account=login_account, begin=begin, end=end,Duplex='',excel_num=excel_report[9])
                                elif excel_report[0] in two_report_list:
                                    if excel_report[0] in('报表-球类报表-子查询（根据其球类查询盘口）','报表-球类报表-子查询-（根据其球类查询盘口）-投注时间','报表-球类报表-子查询-（根据其球类查询盘口）-赛事时间','报表-球类报表-子查询-（根据其球类查询盘口）-结算时间'):
                                        yy_num = self.wt.Company_winlose(agent_id='', member_id='', sportId=sort_num[i][j][-3],marketId=sort_num[i][j][-7], tournamentId='',matchId='',login_account=login_account, begin=begin, end=end,Duplex='',excel_num=excel_report[9])
                                    else:
                                        yy_num = self.wt.Company_winlose(agent_id='', member_id='', sportId=sort_num[i][j][-3],marketId=sort_num[i][j][-7], tournamentId='',matchId=str(sportId_sql_list[i]),login_account=login_account, begin=begin, end=end,Duplex='',excel_num=excel_report[9])
                            elif sort_num[i][j][g] == "总佣金":
                                if excel_report[0] in interface_list:
                                    yy_num = self.wt.total_commission(agent_id=sort_num[i][j][1],member_id='', sportId='',marketId='', tournamentId='', matchId='',login_account=login_account, begin=begin, end=end,Duplex='',ZD='',excel_num=excel_report[9])
                                elif excel_report[0] in two_report_list:
                                    if excel_report[0] in('报表-球类报表-子查询（根据其球类查询盘口）','报表-球类报表-子查询-（根据其球类查询盘口）-投注时间','报表-球类报表-子查询-（根据其球类查询盘口）-赛事时间','报表-球类报表-子查询-（根据其球类查询盘口）-结算时间'):
                                        yy_num = self.wt.total_commission(agent_id='', member_id='',sportId=sort_num[i][j][-3], marketId=sort_num[i][j][-7],tournamentId='', matchId='',login_account=login_account, begin=begin, end=end,Duplex='',ZD='',excel_num=excel_report[9])
                                    else:
                                        yy_num = self.wt.total_commission(agent_id='', member_id='',sportId=sort_num[i][j][-3], marketId=sort_num[i][j][-7],tournamentId='', matchId=str(sportId_sql_list[i]),login_account=login_account, begin=begin,end=end,Duplex='',ZD='',excel_num=excel_report[9])
                            elif type(sort_num[i][j][g])==type(None):
                                yy_num = sort_num[i][j][g]
                            else:
                                yy_num = str(sort_num[i][j][g])
                        else:
                            yy_num = float(sort_num[i][j][g])
                        sql_dict[sql_name_list[g]]=yy_num
                    sql_list.append(sql_dict)
            print("sport_list组装数据对比：", sport_list[0])
            print("sql_list组装数据对比：", sql_list[2])
            # exit(print(sport_list[10],sql_list[10]))
        if mix[0]==str('2-2'):
            # print(sort_num[0][0][0])
            # print(sort_num)
            print(sort_num[0])
            print(sort_num[0][0])
            print(sort_num[0][0][0])
            print(len(sport_list),len(sort_num), len(options_list))
            for i in range(0, len(sort_num)):
                for j in range(0, len(sort_num[i])):
                    for l in range(0, len(sort_num[i][j])):
                        sql_dict = {}
                        for g in range(0,len(sort_num[i][j][l])):
                            if str(type(sort_num[i][j][l][g])) in type_list:
                                if sql_name_list[g] in int_list:
                                    yy_num = int(sort_num[i][j][l][g])
                                elif sort_num[i][j][l][g] == "公司总计":
                                    if excel_report[0] in two_report_list:
                                        yy_num = self.wt.Company_winlose(agent_id='', member_id='',sportId=sort_num[i][j][l][-3], marketId=sort_num[i][j][l][-7],tournamentId='', matchId='',login_account=login_account, begin=begin,end=end,Duplex='',excel_num=excel_report[9])
                                    elif excel_report[0] in ("报表-联赛报表-主查询","报表-联赛报表-主查询-投注时间","报表-联赛报表-主查询-赛事时间","报表-联赛报表-主查询-结算时间"):
                                        yy_num = self.wt.Company_winlose(agent_id='', member_id='', sportId='',marketId='', tournamentId=sort_num[i][j][l][-2],matchId='', login_account=login_account,begin=begin, end=end,Duplex='',excel_num=excel_report[9])
                                    elif excel_report[0] in ("报表-赛事盈亏-主查询","报表-赛事盈亏-主查询-投注时间","报表-赛事盈亏-主查询-赛事时间","报表-赛事盈亏-主查询-结算时间"):
                                        yy_num = self.wt.Company_winlose(agent_id='', member_id='', sportId='',marketId='', tournamentId='',matchId=str(sort_num[i][j][l][-6]),login_account=login_account, begin=begin,end=end,Duplex='',excel_num=excel_report[9])
                                elif sort_num[i][j][l][g] == "总佣金":
                                    if excel_report[0] in two_report_list:
                                        yy_num = self.wt.total_commission(agent_id='', member_id='',sportId=sort_num[i][j][l][-3], marketId=sort_num[i][j][l][-7],tournamentId='', matchId='',login_account=login_account, begin=begin,end=end,Duplex='',ZD='',excel_num=excel_report[9])
                                    elif excel_report[0] in ("报表-联赛报表-主查询","报表-联赛报表-主查询-投注时间","报表-联赛报表-主查询-赛事时间","报表-联赛报表-主查询-结算时间"):
                                        yy_num = self.wt.total_commission(agent_id='', member_id='', sportId='',marketId='', tournamentId=sort_num[i][j][l][-2],matchId='', login_account=login_account,begin=begin, end=end,Duplex='',ZD='',excel_num=excel_report[9])
                                    elif excel_report[0] in ("报表-赛事盈亏-主查询","报表-赛事盈亏-主查询-投注时间","报表-赛事盈亏-主查询-赛事时间","报表-赛事盈亏-主查询-结算时间"):
                                        yy_num = self.wt.total_commission(agent_id='', member_id='', sportId='',marketId='', tournamentId='',matchId=str(sort_num[i][j][l][-6]),login_account=login_account, begin=begin,end=end,Duplex='',ZD='',excel_num=excel_report[9])
                                # elif sort_num[i][j][l][g] == "odds":
                                #     if excel_report[0] == '报表-球类报表-订单查询（根据其盘口查询订单）':
                                #         yy_num = self.sr.credit_odds(order_no=sort_num[i][j][35], bet_type="",AB_list=AB_list, dict=dict)
                                # elif sort_num[i][j][l][g] == "odds_type":
                                #     if excel_report[0] == '报表-球类报表-订单查询（根据其盘口查询订单）':
                                #         yy_num = str(self.sr.total_odds(order_no=sort_num[i][j][35], number=' '))
                                elif type(sort_num[i][j][l][g]) == type(None):
                                    yy_num = sort_num[i][j][l][g]
                                else:
                                    yy_num = str(sort_num[i][j][l][g])
                            else:
                                yy_num = float(sort_num[i][j][l][g])
                            sql_dict[sql_name_list[g]] = yy_num
                        sql_list.append(sql_dict)
            if len(sport_list)>1:
                print("sport_list组装数据对比：",sport_list[1])
                print("sql_list组装数据对比：",sql_list[1])
            else:
                print("sport_list组装数据对比：", sport_list)
                print("sql_list组装数据对比：", sql_list)
        yyts=self.if_error(A=sport_list, B=sql_list,correct_list=correct_list)
        # yyts='调试'
        correct_list=yyts[1]
        if yyts[0]=='错误':
            sport_error.append(f"接口数据值：{sport_list}")
            sport_error.append("--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------")
            sport_error.append(f"SQL数据值：{sql_list}")
            if len(yyds01)==1:
                self.report_list_Compared01(sport_list=sport_list, sql_list=sql_list, Compared=Compared,excel_report=excel_report,begin=begin,end=end,login_account=login_account,rold_id=rold_id)
            elif len(yyds01)==2:
                self.report_list_Compared02(sport_list=sport_list, sql_list=sql_list, Compared=Compared,excel_report=excel_report,begin=begin,end=end,login_account=login_account,rold_id=rold_id)
            else:
                self.report_list_Compared03(sport_list=sport_list, sql_list=sql_list, Compared=Compared,excel_report=excel_report,begin=begin,end=end,login_account=login_account,rold_id=rold_id)
        else:
            pass

        if len(sport_list) == len(correct_list):
            print(f"\033[32m代理{login_account}的{report_name}，接口数据比对全部正确：{len(sport_list)}-{len(correct_list)}\033[0m")
            Prepare_enough_list.append(f"\033[32m代理{login_account}的{report_name}，接口数据比对全部正确\n\n\033[0m")
            sport_error.clear()
        else:
            print(f"\033[31m代理{login_account}的{report_name}，数据错误，需要比对,{len(sport_list)},{len(correct_list)}\033[0m")
            Prepare_enough_list.append(f"\033[31m代理{login_account}的{report_name}，数据错误，需要比对,{len(sport_list)},{len(correct_list)}\033[0m")
            for ff in range(0,len(correct_list)):
                if correct_list[ff] in sql_list:
                    del sql_list[sql_list.index(correct_list[ff])]
            # print(sql_list)
        if yyts[0]=='错误':
            print(f"\033[33m登{rold_id}-({login_account})-{excel_report[0]}-{excel_report[1]}：数据对比结束----------------------------------------------------------------------------------------------\n\n\033[0m")
        else:
            print("\n","\n")
            # print(f"\033[34m部分数据：{sport_error}\033[0m")
            # print(f"\033[34m全部数据：{sport_all_error}\033[0m")

    def time(self, type):
        if type == 1:
            type_time = '投注时间'
        elif type == 2:
            type_time = '赛事时间'
        else:
            type_time = '结算时间'
        return type_time

class report_data(object):
    ROBOT_LIBRARY_SCOPE = 'GLOBAL'
    def __init__(self, mysql_info, mongo_info, *args):
        """
        模拟ctrl给我司推送数据
        """
        # super().__init__(self, mysql_info, mongo_info, *args) #调用基类的—__init__方法
        # self.dbq = DbQuery(mongo_info)
        # self.cf = CommonFunc()
        # self.mysql = MysqlCommonQuery(mysql_info)
        self.my = MysqlFunc(mysql_info)
        self.session = requests.session()
        self.tt=BetController(mysql_info, mongo_info)

    # rsa加密账号、密码
    def rsa_encrypt(self,data):
        '''
        rsa加密（encrypt）
        :param data:
        :return:
        '''
        msg = data.encode('utf-8')
        pub_key = "-----BEGIN PUBLIC KEY-----\nMFwwDQYJKoZIhvcNAQEBBQADSwAwSAJBAL1XuLmIZttk13hmAGVuXiKSfQggfVck" \
                  "p+iNr9jBIxkmBBfmygJ9D5A7lhUbhBEY1SqyGNIHI1DsNLfxfRvW2EcCAwEAAQ==\n-----END PUBLIC KEY-----"
        rsa_key = RSA.importKey(pub_key)
        cipher = Cipher_pkcs1_v1_5.new(rsa_key)
        cipher_text = base64.b64encode(cipher.encrypt(msg)).decode('utf-8')
        # print(cipher_text)
        return cipher_text


    def reading_Excel(self,excel,save_excel,begin,end,URL,data):
        """
        @获循环所有表，取Excel的数据，并把数据写入list中
        """
        global num_count,count
        # 获取 工作簿对象
        workbook = openpyxl.load_workbook(excel[0])
        print("读取表格成功")
        # shenames = workbook.get_sheet_names()
        shenames = workbook.sheetnames
        print("整个表格所有表名:", shenames)  # ['各省市', '测试表']

        #统计用例通过的数量
        execute=0
        #统计用例不通过的数量
        Failed=0
        # 获取用例执行前的时间
        time01 = datetime.datetime.strptime(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),'%Y-%m-%d %H:%M:%S')
        #获取不通过的用例名称
        excel_Failed=[]
        #登入账号
        num_count=0
        count = 0

        #循环Excel所有表
        #(0,12) 报表登0~登3
        #(12,16)盈亏详情
        for gg in range(12,16):
            gg = shenames[gg]
        # for gg in shenames:
            # 获取该表相应的行数和列数
            worksheet= workbook[shenames[shenames.index(gg)]]
            #循环获取列数
            rows=0
            for yy in range(1,1000):
                if worksheet.cell(yy, 1).value!=None:
                    rows=rows+1
                else:
                    break
            # rows = worksheet.max_row
            columns = worksheet.max_column
            print(f"所在\033[32m（{gg}）\033[0m组成有： {rows}行  {columns}列")
            #判断表中第二行数据，是否为空，为空跳过，循环其他表
            content_A2 = worksheet.cell(2, 1).value
            if content_A2!=None:
                # 循环行数
                qqt=100
                for i in range(4,5):
                # for i in range(2, rows+1):
                    if i==qqt:
                        pass
                    else:
                        excel_report = []
                        # 循环列数
                        for j in range(2, int(columns)-2):
                            report=worksheet.cell(i,j).value
                            #获取到的数据写入列表
                            excel_report.append(report)
                        #调用登录接口，获取token，访问接口，获取数据
                        print(excel_report)
                        if print_num==1:
                            print(excel_report)
                            exit()
                        if num_count==excel_report[4]:
                            count=1
                        else:
                            count=0
                            num_count=excel_report[4]
                        self.login(URL=URL,excel_report=excel_report,data=data,count=count,num_count=num_count,begin=begin,end=end)
                        self.Excel_method(URL=URL,token=token,excel_report=excel_report,begin=begin,end=end,parentId=id)

                        '''
                        #调试SQL数据
                        # print(excel_report)
                        # sport_report_list={'account': 'd0d1d2d38y/fceshi0224', 'betAmount': 10.0, 'betIp': '192.168.10.120', 'betIpAddress': '局域网', 'betResult': '输', 'betTime': '2022-06-24 08:57:04', 'betType': '单注', 'level0Commission': 0.0, 'level0CommissionRatio': 0.0, 'level0Percentage': 0.2, 'level0Total': 2.0, 'level0WinOrLose': 2.0, 'level1Commission': 0.0, 'level1CommissionRatio': 0.0, 'level1Percentage': 0.2, 'level1Total': 2.0, 'level1WinOrLose': 2.0, 'level2Commission': 0.0, 'level2CommissionRatio': 0.0, 'level2Percentage': 0.2, 'level2Total': 2.0, 'level2WinOrLose': 2.0, 'level3Commission': 0.0, 'level3CommissionRatio': 0.0, 'level3Percentage': 0.2, 'level3Total': 2.0, 'level3WinOrLose': 2.0, 'memberCommission': 0.0, 'memberCommissionRatio': 0.0, 'memberTotal': -10.0, 'memberWinOrLose': -10.0, 'name': '杜鑫test账号iy', 'odds': 1.69, 'oddsType': '1', 'options': [{'awayTeamName': '芹苴', 'betScore': None, 'homeTeamName': 'QNK广南足球俱乐部', 'marketName': '独赢', 'matchTime': '2022-06-25 06:00:00', 'matchType': '早盘', 'odds': 1.69, 'oddsType': '1', 'orderNo': 'XH4ydmPbRncK', 'outcomeName': 'QNK广南足球俱乐部', 'specifier': '', 'tournamentName': '越南职业足球乙级联赛'}], 'orderNo': 'XH4ydmPbRncK', 'settlementTime': '2022-06-25 08:01:45', 'sportId': 'sr:sport:1', 'sportType': '足球', 'validAmount': 10.0, 'winOrLose': -10.0}
                        # sport_report_dict=['account', 'betAmount', 'betIp', 'betIpAddress', 'betResult', 'betTime', 'betType', 'level0Commission', 'level0CommissionRatio', 'level0Percentage', 'level0Total', 'level0WinOrLose', 'level1Commission', 'level1CommissionRatio', 'level1Percentage', 'level1Total', 'level1WinOrLose', 'level2Commission', 'level2CommissionRatio', 'level2Percentage', 'level2Total', 'level2WinOrLose', 'level3Commission', 'level3CommissionRatio', 'level3Percentage', 'level3Total', 'level3WinOrLose', 'memberCommission', 'memberCommissionRatio', 'memberTotal', 'memberWinOrLose', 'name', 'odds', 'oddsType', 'options', 'orderNo', 'settlementTime', 'sportId', 'sportType', 'validAmount', 'winOrLose']
                        # self.tt.sport_report_sql(begin=begin, end=end, excel_report=excel_report, sport_report_dict=sport_report_list, sport_report_list=sport_report_dict)
                        '''

                        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        #判断用例是否通过
                        if sport_error==[]:
                            execute=execute+1
                            worksheet.cell(i,15,("测试通过"+"\n"+now))
                            worksheet.cell(i,16,'')
                            worksheet.cell(i,17, '')
                            worksheet.cell(i, 18, str(Prepare_enough_list))
                            # if worksheet.cell(i, 15).value==None:
                            #     worksheet.cell(i, 15, str(Prepare_enough_list))
                            # else:
                            #     worksheet.cell(i, 15, "")
                        else:
                            Failed=Failed+1
                            excel_Failed.append(str(excel_report[4]+'-')+(excel_report[0]))
                            print(f"未通过用例:{excel_Failed}")
                            worksheet.cell(i,15, ("测试不通过"+"\n"+now))
                            worksheet.cell(i,16, str(sport_all_error))
                            worksheet.cell(i,17, str(sport_error))
                            worksheet.cell(i,18, str(Prepare_enough_list))
                        workbook.save(filename=save_excel[0])
            else:
                continue
        time02=datetime.datetime.strptime(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),'%Y-%m-%d %H:%M:%S')
        date_time=time02-time01
        print(f"\n\n开始时间:{time01}~结束时间:{time02}")
        print(f"执行通过用例共计\033[32m{execute}条\033[0m,未通过用例共计\033[31m{Failed}条\033[0m,共计消耗时间{date_time}")
        if excel_Failed!=[]:
            print(f"执行未通过的用例名称：{excel_Failed}")

        workbook.close()

    def bf_request(self,url,headers,data,excel_report,begin,end):
        for loop in range(0,5):
            try:
                session = requests.session()
                #判断请求方式
                request_method=['post']
                if excel_report[10]==request_method[0]:
                    response = session.post(url=url, headers=headers, json=data)
                    # print(url,headers,data)
                else:
                    response = session.get(url=url, headers=headers, params=data)
                # #判断接口返回code,是否为：200
                interface_returns_data=response.json()
                # if interface_returns_data['message']=="签名处理异常! 可能原因: 系统维护中，请稍后再试" or interface_returns_data['message']=='登录状态已过期，请重新登录':
                #     print(f"第{loop}次重连，获取token")
                #     time.sleep(loop)
                #     self.login(URL=URL,excel_report=excel_report,data=0,count=0,num_count=num_count,begin=begin,end=end,TOKEN=1)
                #     break

                if response.status_code != 200:
                    print(f'{response}请求超时:{interface_returns_data}')
                    return response
                else:
                    # print(f"接口{url}，请求{data}成功")
                    return response
            #请求异常判断
            except ConnectionError:
                time.sleep(2)
                continue
            #其他报错信息
            except Exception as e:
              print(f"\33[31m当前接口接口调用失败，请求检查接口,失败信息：{e}\33[0m")


    def login(self,URL,excel_report,data,count,num_count,begin,end):
        """
        @第一次获取token，并保持token连接中，便于后面其他接口访问
        """
        global token,id

        if data==0:
            print(count,num_count)
            # 登0~登3登录获取token
            login_account = excel_report[4]
            if count == 0:
                session = requests.session()

                url = str(URL) + "/system/accountLogin"
                headers = {
                    'Content-Type': 'application/json;charset=UTF-8'
                }
                #Bfty123456,Bf123456
                #Abc123456,qq123456
                data = {"userName": self.rsa_encrypt(data=login_account),
                        "password": self.rsa_encrypt(data=pw),
                        "securityCode": pwa,
                        "loginDiv": 555666}
                # data = {"userName": self.rsa_encrypt(data=login_account),
                #         "password": self.rsa_encrypt(data="Bfty123456"),
                #         "securityCode": "Bf123456",
                #         "loginDiv": 555666}
                response = session.post(url=url, headers=headers, json=data)
                # 返回结果json转化
                results = json.loads(response.text)
                if results['code'] == 0:
                    pass
                    # print("登录成功"+"\n")
                else:
                    print(results)

                token = results['data']['token']
                id = results['data']['id']


                # 调用过渡函数，保持token登录状态：
                url01 = str(URL) + "/system/userDetailByToken?"
                headers01 = {
                    'Content-Type': 'application/json;charset=UTF-8',
                    'LoginDiv': '555666',
                    'Account_Login_Identify': token
                }
                data01 = {
                    "token": token,
                    "accountId": id
                }
                response01 = session.get(url=url01, headers=headers01, params=data01)
                # 返回结果json转化
                results01=json.loads(response01.text)
                print(f"过渡接口调用token：{results01}")
                time.sleep(3)
                count=count+1
            else:
                token=token
        else:
            token=token
        print(token)
        return token

    #根据Excel类型，去调用对应的接口方法
    def Excel_method(self,URL,token,excel_report,begin,end,parentId):
        print(parentId)
        #调用方法：
        report_yyds = self.DL_user_role_id(excel_report=excel_report)
        if excel_report[3] == 1:
            self.sport_report01(URL=URL, token=token, excel_report=excel_report, begin=begin, end=end,parentId=parentId,login_account=report_yyds[0],rold_id=report_yyds[1])
        elif excel_report[3] == 2:
            self.sport_report02(URL=URL, token=token, excel_report=excel_report, begin=begin, end=end,parentId=parentId,login_account=report_yyds[0],rold_id=report_yyds[1])
        else:
            self.sport_report03(URL=URL, token=token, excel_report=excel_report, begin=begin, end=end,parentId=parentId,login_account=report_yyds[0],rold_id=report_yyds[1])

    #处理token过期方法
    def token_method(self,URL, token, excel_report, begin,end,response,parentId):
        # 判断接口返回code,是否为：200
        interface_returns_data = response.json()
        if interface_returns_data['message'] == "签名处理异常! 可能原因: 系统维护中，请稍后再试" or interface_returns_data['message'] == '登录状态已过期，请重新登录':
            time.sleep(3)
            print("\33[31mtoken已过期，正在获取\33[0m")
            self.login(URL=URL, excel_report=excel_report, data=0, count=0, num_count=num_count, begin=begin, end=end)
            time.sleep(5)
            self.Excel_method(URL=URL,token=token, excel_report=excel_report, begin=begin,end=end,parentId=parentId)
        else:
            pass
            # print("\33[32mtoken未过期，可以继续使用\33[0m")

    # 处理登0~登2，代理取登3ID查询
    def D3(self, excel_report):
        report_yyds=self.DL_user_role_id(excel_report=excel_report)
        sql01 = f"SELECT c.proxy3_id FROM m_account as a LEFT JOIN o_account_order as c ON c.proxy{report_yyds[1]}_id=a.id WHERE a.login_account='{report_yyds[0]}' GROUP BY c.proxy3_id"
        sort_num01 = self.my.query_data(sql01, db_name='bfty_credit')
        proxy3_id = sort_num01[0][0]
        return proxy3_id

    #判断excel输入登入账号是否正确，等级是否正确
    def DL_user_role_id(self,excel_report):
        # 判断输入登入账号是否正确
        sql01 = f"SELECT login_account FROM m_account WHERE login_account='{excel_report[4]}'"
        sort_num01 = self.my.query_data(sql01, db_name='bfty_credit')
        if sort_num01 == ():
            sql02 = f"SELECT login_account FROM m_account WHERE account='{excel_report[4]}'"
            sort_num02 = self.my.query_data(sql02, db_name='bfty_credit')
            login_account = sort_num02[0][0]
        else:
            login_account=excel_report[4]
        # 判断Excel的代理等级是否正确
        sql03 = f"SELECT role_id FROM m_account WHERE login_account='{login_account}'"
        sort_num03 = self.my.query_data(sql03, db_name='bfty_credit')
        if sort_num03[0][0] == int(excel_report[5]):
            rold_id=excel_report[5]
        else:
            rold_id=sort_num03[0][0]
        return login_account,rold_id


    def sport_report01(self,URL, token, excel_report, begin, end,parentId,login_account,rold_id):
        global sportId_list
        print(f"\33[34m登{rold_id}-({login_account})-正在执行接口数据{excel_report[0]}-{excel_report[1]}\33[0m")

        sport_report_dict = []
        sport_report_list = []
        sportId_list=[]

        url = str(URL)+(excel_report[1])
        headers=eval(excel_report[6])
        if excel_report[0] in sport_report02_list:
            data = eval(excel_report[7])[0]
            if rold_id==3:
                data['parentId'] = parentId
            else:
                parentId=self.D3(excel_report=excel_report)
                data['parentId'] = parentId
            print(data)
        elif excel_report[0] in('首页-余额详细资料-主查询','首页-统计-主查询'):
            data=None
        else:
            data = eval(excel_report[7])[0]
        session = requests.session()
        # print(headers)
        # print(data)

        #获取接口数据，获取失败try3次
        response= self.bf_request(url=url,headers=headers,data=data,excel_report=excel_report,begin=begin,end=end)
        # 判断接口token是否过期
        self.token_method(URL=URL, token=token, excel_report=excel_report, begin=begin, end=end, response=response,parentId=parentId)
        results = json.loads(response.text)
        # print(results)

        if excel_report[0] in module_list:
            yyds = results['data']
            if excel_report[0] =='首页-统计-主查询':
                yyts = {}
                yyts_list = []
                for key, value in yyds['creditsAmount'].items():
                    yyts[key] = yyds['creditsAmount'][key]
                yyts['householdsNumList'] = yyds['householdsNumList']
                yyds = yyts
        elif excel_report[0]=='首页-余额详细资料-主查询':
            yyds=[]
            yyds.append(results['data']['data'])
        elif excel_report[0]=='报表-已取消注单-主查询':
            yyds = results['data']['data']['data']
        else:
            yyds = results['data']['data']


        if yyds==[] or yyds==None:
            pass
        else:
            str_sum=excel_report[2]
            # print(yyds)
            for i in range(0,len(yyds)):
                if excel_report[2] != None:
                    ppxt_lsit = str_sum.split(",")
                    for j in ppxt_lsit:
                        # print(j)
                        del yyds[i][j]
                sport_report_list.append(yyds[i])
                # print(yyds[i])
            # print(sport_report_list)
            for key,value in yyds[0].items():
                sport_report_dict.append(key)

        time_list=[" 00:00:00"," 23:59:59"]
        #判断接口数据为空时，不打印信息
        if sport_report_list==[]:
            print(f"\033[32m接口返回信息为空\33[0m")
        else:
            print(sport_report_dict)
            if len(sport_report_list)>1:
                print(sport_report_list[1])
            else:
                print(sport_report_list)
        self.tt.sport_report_sql(begin=begin+time_list[0], end=end+time_list[1],excel_report=excel_report,sport_report_dict=sport_report_dict,sport_report_list=sport_report_list,login_account=login_account,rold_id=rold_id)
        return sport_report_dict


    def sport_report02(self,URL, token, excel_report, begin, end,parentId,login_account,rold_id):
        global sportId_list,sportId_yyds_lsit
        print(f"\33[34m登{rold_id}-({login_account})-正在执行接口数据{excel_report[0]}-{excel_report[1]}\33[0m")

        sportId_list = []
        marketID_list=[]
        matchId_id_list=[]
        sport_report_list=[]
        sport_report_dict = []
        sportId_yyds_lsit = []
        yyds=''
        index=''
        for url in eval(excel_report[1]):
            headers = eval(excel_report[6])
            if eval(excel_report[1]).index(url) == 0:
                url = str(URL) + url
                data = eval(excel_report[7])[0]
                if  excel_report[0] in ("盈亏详情-登0-登3-会员-查看订单详情-子查询",'盈亏详情-登1-登3-会员-查看订单详情-子查询','盈亏详情-登2-登3-会员-查看订单详情-子查询'):
                    parentId=self.D3(excel_report=excel_report)
                    data['parentId'] = parentId
                else:
                    data['parentId'] = parentId
                # 获取接口数据，获取失败try3次
                response = self.bf_request(url=url,headers=headers,data=data,excel_report=excel_report,begin=begin,end=end)
                # 判断接口token是否过期
                self.token_method(URL=URL, token=token, excel_report=excel_report, begin=begin, end=end,response=response,parentId=parentId)
                results = json.loads(response.text)


                if excel_report[0] in module_list:
                    yyds = results['data']
                else:
                    yyds = results['data']['data']

                if yyds == [] or yyds==None:
                    pass
                else:
                    for i in yyds:
                        if i[excel_report[8]]=='串关':
                            pass
                        else:
                            sportId_list.append(i[excel_report[8]])
            else:
                if yyds == []:
                    break
                else:
                    #判断接口数量是否大于2，或小于2
                    if len(eval(excel_report[1]))>2:
                        index=2
                        if eval(excel_report[1]).index(url)==1:
                            url = str(URL) + url
                            market_id_list = []
                            if excel_report[0] in('报表-球类报表-订单查询（根据其盘口查询订单）','报表-联赛报表-订单查询（根据其盘口查询订单）','报表-赛事盈亏-订单查询（根据其盘口查询订单）','报表-混合串关-订单查询（根据其会员查询订单）','报表-已取消注单-主查询'):
                                for sportId in sportId_list:
                                    data = eval(excel_report[7])[1]
                                    data[excel_report[8]] = sportId
                                    # 获取接口数据，获取失败try3次
                                    response = self.bf_request(url=url,headers=headers,data=data,excel_report=excel_report,begin=begin,end=end)
                                    # 判断接口token是否过期
                                    self.token_method(URL=URL, token=token, excel_report=excel_report, begin=begin, end=end, response=response)
                                    results = json.loads(response.text)
                                    yyds = results['data']['data']

                                    if yyds == [] or yyds==None:
                                        pass
                                    else:
                                        for i in yyds:
                                            if excel_report[0]=="报表-球类报表-订单查询（根据其盘口查询订单）":
                                                market_id_list.append(i['marketId'])
                                            elif excel_report[0]=="报表-联赛报表-订单查询（根据其盘口查询订单）":
                                                market_id_list.append(i['tournamentId'])
                                            elif excel_report[0]=="报表-赛事盈亏-订单查询（根据其盘口查询订单）":
                                                market_id_list.append(i['matchId'])
                                    marketID_list.append(market_id_list)
                        else:
                            if eval(excel_report[1]).index(url)==2:
                                url = str(URL) + url
                                for sportId in sportId_list:
                                    for marketID in marketID_list[sportId_list.index(sportId)]:
                                        data = eval(excel_report[7])[2]
                                        data[excel_report[8]] = sportId
                                        if excel_report[0]=='报表-球类报表-订单查询（根据其盘口查询订单）':
                                            data['marketId'] = marketID
                                        elif excel_report[0] == "报表-联赛报表-订单查询（根据其盘口查询订单）":
                                            data['tournamentId'] = marketID
                                        elif excel_report[0]=='报表-赛事盈亏-订单查询（根据其盘口查询订单）':
                                            data['matchId'] = marketID
                                        # 获取接口数据，获取失败try3次
                                        response = self.bf_request(url=url,headers=headers,data=data,excel_report=excel_report,begin=begin,end=end)
                                        # 判断接口token是否过期
                                        self.token_method(URL=URL, token=token, excel_report=excel_report, begin=begin,end=end, response=response)
                                        results = json.loads(response.text)
                                        if excel_report[0]=='报表-赛事盈亏-订单查询（根据其盘口查询订单）':
                                            yyds = results['data']['data']
                                            for i in yyds:
                                                matchId_id_list.append(i['marketId'])
                                        else:
                                            yyds = results['data']['data']['data']
                                            sportId_yyds_lsit.append(yyds)
                            else:
                                if excel_report[0]=='报表-赛事盈亏-订单查询（根据其盘口查询订单）':
                                    url = str(URL) + url
                                    for marketID in matchId_id_list:
                                        data = eval(excel_report[7])[3]
                                        data['marketId'] = marketID
                                        # 获取接口数据，获取失败try3次
                                        response = self.bf_request(url=url,headers=headers,data=data,excel_report=excel_report,begin=begin,end=end)
                                        # 判断接口token是否过期
                                        self.token_method(URL=URL, token=token, excel_report=excel_report, begin=begin,end=end, response=response)
                                        results = json.loads(response.text)
                                        yyds = results['data']['data']['data']
                                        sportId_yyds_lsit.append(yyds)
                    else:
                        url = str(URL) + url
                        for sportId in sportId_list:
                            data = eval(excel_report[7])[1]
                            if excel_report[0]in ("未完成交易-登0-登3-会员-查看订单详情-子查询","盈亏详情-登0-登3-会员-查看订单详情-子查询",'盈亏详情-登1-登3-会员-查看订单详情-子查询','盈亏详情-登2-登3-会员-查看订单详情-子查询','盈亏详情-登3-会员-查看订单详情-子查询'):
                                data['parentId']=sportId
                            elif excel_report[0]=='报表-混合串关-订单查询（根据其会员查询订单）':
                                data['account'] = sportId
                            else:
                                data[excel_report[8]] = sportId
                            # 获取接口数据，获取失败try3次
                            response = self.bf_request(url=url,headers=headers,data=data,excel_report=excel_report,begin=begin,end=end)
                            # 判断接口token是否过期
                            self.token_method(URL=URL, token=token, excel_report=excel_report, begin=begin, end=end,response=response,parentId=parentId)
                            results = json.loads(response.text)

                            if excel_report[0] in module_list:
                                if excel_report[0] in ("未完成交易-登0-登3-会员-查看订单详情-子查询","盈亏详情-登0-登3-会员-查看订单详情-子查询",'盈亏详情-登1-登3-会员-查看订单详情-子查询','盈亏详情-登2-登3-会员-查看订单详情-子查询','盈亏详情-登3-会员-查看订单详情-子查询'):
                                    if results['data']['data'] == []:
                                        pass
                                    else:
                                        # print(results)
                                        #data字典里的数据
                                        yyds = results['data']['data']
                                        if excel_report[0]=="未完成交易-登0-登3-会员-查看订单详情-子查询":
                                            Total_Amount = results['data']['totalData']['betAmount']
                                            for ppds in range(0,len(yyds)):
                                                data =yyds[ppds]['bettingTime']
                                                data = data.replace("T", " ")
                                                data = data.replace(".000Z", "")
                                                yyds[ppds]['bettingTime']=data
                                                yyds[ppds]['Total_Amount']=Total_Amount
                                        elif excel_report[0]in ("盈亏详情-登0-登3-会员-查看订单详情-子查询",'盈亏详情-登1-登3-会员-查看订单详情-子查询','盈亏详情-登2-登3-会员-查看订单详情-子查询','盈亏详情-登3-会员-查看订单详情-子查询'):
                                            # 双字典totalData的数据
                                            totalData = ['level0Commission', 'level0Total', 'level0WinOrLose','level1Commission', 'level1Total', 'level1WinOrLose','level2Commission', 'level2Total', 'level2WinOrLose','level3Commission', 'level3Total', 'level3WinOrLose', 'memberCommission', 'memberTotal', 'memberWinOrLose', 'validAmount','winOrLose']
                                            totalData_dict = {}
                                            # print(results['data']['totalData'])
                                            #把第二字典需要对比的数据取出来，放进第一个data里面，方便对比
                                            for ppds in range(0, len(yyds)):
                                                for total in totalData:
                                                    ppt = " total_" + total
                                                    yyds[ppds][ppt]= results['data']['totalData'][total]
                                            # print(yyds[0])
                                else:
                                    yyds = results['data']
                            else:
                                if excel_report[0]=='报表-混合串关-订单查询（根据其会员查询订单）':
                                    yyds = results['data']['data']['data']
                                else:
                                    yyds = results['data']['data']
                            sportId_yyds_lsit.append(yyds)
                            # print(sportId_yyds_lsit)
                        # print(sportId_yyds_lsit)
        if yyds==[] or  yyds==None:
            pass
        else:
            report02_list = []
            for gg in range(0, len(sportId_yyds_lsit)):
                for yy in range(0, len(sportId_yyds_lsit[gg])):
                    report02_list.append(sportId_yyds_lsit[gg][yy])
            # print(report02_list)

            str_sum = excel_report[2]
            for i in range(0, len(report02_list)):
                if excel_report[0] in ("未完成交易-登0-登3-会员-查看订单详情-子查询","盈亏详情-登0-登3-会员-查看订单详情-子查询",'盈亏详情-登1-登3-会员-查看订单详情-子查询','盈亏详情-登2-登3-会员-查看订单详情-子查询','盈亏详情-登3-会员-查看订单详情-子查询'):
                    print(report02_list)
                    for options in report02_list[i]['options']:
                        del options['matchTime']
                if excel_report[2] != None:
                    ppxt_lsit = str_sum.split(",")
                    # print(ppxt_lsit)
                    for j in ppxt_lsit:
                        # print(sportId_yyds_lsit[i])
                        del report02_list[i][j]
                sport_report_list.append(report02_list[i])

            # print(sportId_yyds_lsit[0])
            # exit(print(sportId_yyds_lsit))
            for key, value in sportId_yyds_lsit[0][0].items():
                sport_report_dict.append(key)

        time_list = [" 00:00:00", " 23:59:59"]
        # 判断接口数据为空时，不打印信息
        if sport_report_list == []:
            print(f"\033[32m接口返回信息为空\33[0m")
        else:
            print(sport_report_dict)
            print(sport_report_list[1])
        self.tt.sport_report_sql(begin=begin + time_list[0], end=end + time_list[1], excel_report=excel_report,sport_report_dict=sport_report_dict, sport_report_list=sport_report_list,login_account=login_account,rold_id=rold_id)
        return sportId_list

    def sport_report03(self, URL, token, excel_report, begin, end,parentId,login_account,rold_id):
        global sportId_list, sportId_yyds_lsit
        print(f"\33[34m登{rold_id}-({login_account})-正在执行接口数据{excel_report[0]}-{excel_report[1]}\33[0m")

        sportId_list = []
        sportId_yyds_lsit=[]
        sport_report_list = []
        sport_report_dict = []
        yyds = ''
        #判断代理等级
        if int(rold_id)==0:
            rold=3
        elif int(rold_id)==1:
            rold=2
        elif int(rold_id)==2:
            rold=1

        for yytt in range(0,rold):
            headers = eval(excel_report[6])
            if yytt==0:
                url = str(URL) + excel_report[1]
                data = eval(excel_report[7])[0]
                data['parentId']=parentId
                print(data)
                # 获取接口数据，获取失败try3次
                response = self.bf_request(url=url,headers=headers,data=data,excel_report=excel_report,begin=begin,end=end)
                # 判断接口token是否过期
                self.token_method(URL=URL, token=token, excel_report=excel_report, begin=begin, end=end,response=response,parentId=parentId)
                results = json.loads(response.text)

                if excel_report[0] in module_list:
                    yyds = results['data']
                else:
                    yyds = results['data']['data']

                sportId_yyds_lsit.append(yyds)

                if yyds == [] or yyds==None:
                    pass
                else:
                    for i in yyds:
                        if i[excel_report[8]] == '串关':
                            pass
                        elif excel_report[0] in interface_list:
                            sportId_list.append(i[excel_report[8]])
                            sportId_list.append(i['accountId'])
                            roid_list.append(i['levelId'])
                        else:
                            sportId_list.append(i[excel_report[8]])
            else:
                if yyds == []:
                    break
                else:
                    url = str(URL) + excel_report[1]
                    #特殊处理（未完成交易-登0-主查询接口的bady传参）
                    if excel_report[0] in interface_list:
                        data = eval(excel_report[7])[0]
                        # print(excel_report[8],sportId_list[yytt])
                        data[excel_report[8]]=sportId_list[yytt]
                    else:
                        data = eval(excel_report[7])[0]
                    # 获取接口数据，获取失败try3次
                    response = self.bf_request(url=url,headers=headers,data=data,excel_report=excel_report,begin=begin,end=end)
                    # 判断接口token是否过期
                    self.token_method(URL=URL, token=token, excel_report=excel_report, begin=begin, end=end,response=response,parentId=parentId)
                    results = json.loads(response.text)



                    if excel_report[0] in module_list:
                        yyds = results['data']
                    else:
                        yyds = results['data']['data']
                    sportId_yyds_lsit.append(yyds)

                    for i in yyds:
                        if i[excel_report[8]] == '串关':
                            pass
                        elif excel_report[0] in interface_list:
                            sportId_list.append(i['accountId'])
                            roid_list.append(i['levelId'])
                            # print(sportId_list, roid_list)
                        else:
                            sportId_list.append(i[excel_report[8]])

        if yyds == []:
            pass
        else:
            report03_list = []
            for gg in range(0, len(sportId_yyds_lsit)):
                for yy in range(0, len(sportId_yyds_lsit[gg])):
                    report03_list.append(sportId_yyds_lsit[gg][yy])
            # print(report02_list)


            str_sum = excel_report[2]
            for i in range(0, len(report03_list)):
                if excel_report[2] != None:
                    ppxt_lsit = str_sum.split(",")
                    # print(ppxt_lsit)
                    for j in ppxt_lsit:
                        # print(sportId_yyds_lsit[i])
                        del report03_list[i][j]
                sport_report_list.append(report03_list[i])

            for key, value in sportId_yyds_lsit[0][0].items():
                sport_report_dict.append(key)


        # 判断接口数据为空时，不打印信息
        if sport_report_list == []:
            print(f"\033[32m接口返回信息为空\33[0m")
        else:
            print(sport_report_dict)
            if len(sport_report_list)>1:
                print(sport_report_list[1])
            else:
                print(sport_report_list)
        time_list = [" 00:00:00", " 23:59:59"]
        self.tt.sport_report_sql(begin=begin + time_list[0], end=end + time_list[1], excel_report=excel_report,sport_report_dict=sport_report_dict, sport_report_list=sport_report_list,login_account=login_account,rold_id=rold_id)
        return sportId_list

if __name__ == "__main__":
    # MDE环境
    URL = "https://mdesearch.betf.best"

    #120环境
    # mongo_inf = ['app', '123456', '192.168.10.120', '27017']
    # mysql_inf = ['192.168.10.121', 'root', 's3CDfgfbFZcFEaczstX1VQrdfRFEaXTc', '3306']
    #MDE环境
    mongo_inf = ['sport_test', 'BB#gCmqf3gTO5777', '35.194.233.30', '27017']
    mysql_inf = ['35.194.233.30', 'root', 'BB#gCmqf3gTO5b*', '3306']

    AB_list = ["A", "B", "C", "D", "E", "F", ]
    dict = {'A': [], 'B': [], 'C': [], 'D': [], 'E': [], 'F': []}


    # mf = MongoFunc(mongo_inf)
    # bc =BetController(mysql_inf, mongo_inf)

    cc=report_data(mysql_inf,mongo_inf)
    for pkk in range(0,1):
        #参数化接口数据列表
        # sport_report_list = []
        # 参数化代理账号数据列表
        Dl_list=["d0","d10","d2","d3"]
        # 参数化EXCEL路径列表
        # excel = [(r"C:\\test\\d0_comparison_report.xlsx"),(r"C:\\test\\d0_report.xlsx")]
        excel=[r"C:\\test\\d0_report.xlsx"]
        # save_excel = ["C:\\test\\d0_comparison_report.xlsx"]
        save_excel=["C:\\test\\d0_report.xlsx"]
        # 参数化接口返回数据，字典取值判断list，在list中，取results['data'],不在列表中，取results['data']['data']
        module_list = ["总投注-混合串关-主查询","总投注-混合串关-子查询(查询其注单号，包含的比赛)",
                       "未完成交易-登0-主查询","未完成交易-登0-登3-会员-子查询","未完成交易-登0-登3-会员-查看订单详情-子查询",
                       "盈亏详情-登0-主查询","盈亏详情-登0-登3-会员-子查询","盈亏详情-登0-登3-会员-查看订单详情-子查询","首页-统计-主查询",
                       "盈亏详情-登1-主查询", "盈亏详情-登1-登3-会员-子查询", "盈亏详情-登1-登3-会员-查看订单详情-子查询",
                       "盈亏详情-登2-主查询", "盈亏详情-登2-登3-会员-子查询", "盈亏详情-登2-登3-会员-查看订单详情-子查询",
                       "盈亏详情-登3-主查询", "盈亏详情-登3-会员-子查询", "盈亏详情-登3-会员-查看订单详情-子查询"
                       ]
        roid_list=[]
        proxy_id_list=['0','1','2','3']
        int_list=['levelId','numberOfBets',"betType"]
        type_list=["<class 'str'>", "<class 'datetime.datetime'>", "<class 'NoneType'>", "<class 'int'>"]
        sport_report02_list=['盈亏详情-登0-登3-会员-子查询','盈亏详情-登1-登3-会员-子查询','盈亏详情-登2-登3-会员-子查询','盈亏详情-登3-会员-子查询']
        interface_list = ["未完成交易-登0-主查询", "盈亏详情-登0-主查询", "盈亏详情-登1-主查询", "盈亏详情-登2-主查询", "盈亏详情-登3-主查询"]
        options_report_list=['未完成交易-登0-登3-会员-查看订单详情-子查询',
                             '盈亏详情-登0-登3-会员-查看订单详情-子查询','盈亏详情-登0-登3-会员-子查询',
                             '盈亏详情-登1-登3-会员-查看订单详情-子查询','盈亏详情-登1-登3-会员-子查询',
                             '盈亏详情-登2-登3-会员-查看订单详情-子查询','盈亏详情-登2-登3-会员-子查询',
                             '盈亏详情-登3-会员-查看订单详情-子查询','盈亏详情-登3-会员-子查询']

        one_report_list=['报表-球类报表-主查询','报表-联赛报表-主查询','报表-赛事盈亏-主查询','报表-混合串关-主查询','报表-球类报表-主查询-投注时间','报表-联赛报表-主查询-投注时间','报表-赛事盈亏-主查询-投注时间','报表-混合串关-主查询-投注时间','报表-球类报表-主查询-赛事时间','报表-联赛报表-主查询-赛事时间','报表-赛事盈亏-主查询-赛事时间','报表-混合串关-主查询-赛事时间','报表-球类报表-主查询-结算时间','报表-联赛报表-主查询-结算时间','报表-赛事盈亏-主查询-结算时间','报表-混合串关-主查询-结算时间']
        two_report_list = ['报表-球类报表-子查询（根据其球类查询盘口）', '报表-赛事盈亏-子查询（根据其赛事查询其盘口）',
                           '报表-球类报表-子查询-（根据其球类查询盘口）-投注时间', '报表-赛事盈亏-子查询-（根据其赛事查询其盘口）-投注时间',
                           '报表-球类报表-子查询-（根据其球类查询盘口）-赛事时间', '报表-赛事盈亏-子查询-（根据其赛事查询其盘口）-赛事时间',
                           '报表-球类报表-子查询-（根据其球类查询盘口）-结算时间', '报表-赛事盈亏-子查询-（根据其赛事查询其盘口）-结算时间']

        # yyds=cc.login(URL=URL,begin="2022-06-04",end="2022-06-10")
        # ppds=bc.
        # (begin="2022-06-04 00:00:00",end="2022-06-10 23:59:59")

        # time_new_list=[]
        # for i in range(0,100):
        #     time.sleep(600)
        #     time01 = datetime.datetime.strptime(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),'%Y-%m-%d %H:%M:%S')
        #     time_new_list.append(time01)
        #     #调用本地Exce读取数据
        #     ccds=cc.reading_Excel(excel=excel, save_excel=save_excel,begin="2022-06-15",end="2022-06-21", URL=URL)
        # print(time_new_list)

        # 调用本地Exce读取数据
        print_num=0
        # token='eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpZCI6IjE1MzE1MTYwMTc4NDc4Njk0NDIiLCJleHAiOjE2NTg1NDk2ODIsInVzZXJuYW1lIjoiZDAifQ.Vxps5rfymwng_7N8Ikk_c4sSAew6OaMXzfuP1W4PVP0'
        print(excel,save_excel)
        # Bfty123456,Bf123456
        # Abc123456,qq123456
        pw = 'Bfty123456'
        pwa = 'Bf123456'
        # pw ='Abc123456'
        # pwa ='qq123456'
        # pw = 'Bfty123456'
        # pwa = 'Agent0'
        # username='y0'
        # username = 'sh'
        ccds=cc.reading_Excel(excel=excel, save_excel=save_excel,begin="2022-07-20",end="2022-07-25", URL=URL,data=0)


        #调试SQL数据:
        # excel_report=['报表-混合串关-订单查询（根据其会员查询订单）', '"/winOrLost/multiterm","/winOrLost/order/details"', 'companyCommission,companyCommissionRatio,companyPercentage,companyTotal,companyWinOrLose', 2, "{\n        'Content-Type': 'application/json;charset=UTF-8',\n        'LoginDiv': '555666',\n        'Account_Login_Identify': token\n    }", '[{"matchId":"","sportId":"","queryDateType":3,"begin":begin,"end":end,"searchAccount":"","page":1,"limit":500},{"begin":begin,"end":end,"dateType":3,"page":1,"limit":500,"sportId":"","marketId":"串关","account":"fceshi070","queryDateType":3,"tournamentId":""}]', 'loginAccount', 'post', '@f"SELECT CONCAT(c.user_name,\'/\',c.login_account) AS \'账号/登入账号\',any_value (c.bet_amount) \'投注额\',any_value (c.bet_ip),any_value (c.ip_address),CASE any_value (c.mix_num) WHEN \'2_1_0\' THEN \'2串1\' WHEN \'3_1_0\' THEN \'3串1\' WHEN \'4_1_0\' THEN \'4串1\' WHEN \'5_1_0\' THEN \'5串1\' WHEN \'6_1_0\' THEN \'6串1\' WHEN \'7_1_0\' THEN \'7串1\' WHEN \'8_1_0\' THEN \'8串1\' WHEN \'9_1_0\' THEN \'9串1\' WHEN \'10_1_0\' THEN \'10串1\' WHEN \'2_3_0\' THEN \'2串1\' WHEN \'3_4_1\' THEN \'3串4\' WHEN \'3_4_0\' THEN \'3串1\' WHEN \'2_6_0\' THEN \'2串1\' WHEN \'4_11_1\' THEN \'4串11\' WHEN \'4_5_0\' THEN \'4串1\' WHEN \'3_10_0\' THEN \'3串1\' WHEN \'2_10_0\' THEN \'2串1\' WHEN \'5_26_1\' THEN \'5串26\' WHEN \'5_6_0\' THEN \'5串1\' WHEN \'4_15_0\' THEN \'4串1\' WHEN \'3_20_0\' THEN \'3串1\' WHEN \'2_15_0\' THEN \'2串1\' WHEN \'6_57_1\' THEN \'6串57\' END AS \'注单类型\',(CASE any_value (c.settlement_result) WHEN \'1\' THEN \'赢\' WHEN \'2\' THEN \'输\' WHEN \'5\' THEN \'平局走水\' WHEN \'6\' THEN \'注单取消\' END) AS \'注单状态\',any_value (c.create_time),CASE any_value(c.bet_type) WHEN \'1\' THEN \'单注\' WHEN \'2\' THEN \'串关\' WHEN \'3\' THEN \'复式串关\' END \'注单状态\',any_value (c.level0_backwater_amount) \'总代佣金\',c.company_retreat_proportion \'总代佣金比例\',c.level0_actual_percentage \'总代占成数\',any_value (c.level0_win_or_lose) \'总代总计\',(any_value (c.level0_win_or_lose)-any_value (c.level0_backwater_amount)) \'总代输/赢\',any_value (c.level1_backwater_amount) \'一级代理佣金\',c.level0_retreat_proportion \'一级代理佣金比例\',c.level1_actual_percentage \'一级代理占成数\',any_value (c.level1_win_or_lose) \'一级代理总计\',(any_value (c.level1_win_or_lose)-any_value (c.level1_backwater_amount)) \'一级代理输/赢\',any_value (c.level2_backwater_amount) \'二级代理佣金\',c.level1_retreat_proportion \'二级代理佣金比例\',c.level2_actual_percentage \'二级代理占成数\',any_value (c.level2_win_or_lose) \'二级代理总计\',(any_value (c.level2_win_or_lose)-any_value (c.level2_backwater_amount)) \'二级代理输/赢\',any_value (c.level3_backwater_amount) \'三级代理佣金\',c.level2_retreat_proportion \'三级代理佣金比例\',c.level3_actual_percentage \'三级代理占成数\',any_value (c.level3_win_or_lose) \'三级代理总计\',(any_value (c.level3_win_or_lose)-any_value (c.level3_backwater_amount)) \'三级代理输/赢\',any_value (c.backwater_amount) \'会员佣金\',c.level3_retreat_proportion \'会员佣金比例\',(any_value (c.handicap_win_or_lose)+any_value (c.backwater_amount)) \'会员总计\',any_value (c.handicap_win_or_lose) \'会员输/赢\',any_value (c.mix_num),any_value (b.NAME) \'名称\',\'odds\' as \'总赔率\',\'odds_type\'as \'赔率类型\',\'options\' AS \'比赛\',c.order_no,c.award_time,c.sport_id,CASE c.sport_id WHEN \'sr:sport:1\' THEN \'足球\' WHEN \'sr:sport:20\' THEN \'乒乓球\' WHEN \'sr:sport:5\' THEN \'网球\' WHEN \'sr:sport:3\' THEN \'棒球\' WHEN \'sr:sport:2\' THEN \'篮球\' WHEN \'sr:sport:31\' THEN \'羽毛球\' WHEN \'sr:sport:4\' THEN \'冰球\' WHEN \'sr:sport:23\' THEN \'排球\' END AS \'球类\',c.efficient_amount,any_value (c.handicap_win_or_lose) \'会员输/赢\' FROM o_account_order AS c LEFT JOIN u_user AS b ON b.id=c.user_id LEFT JOIN o_account_order_match AS a ON a.order_no=c.order_no WHERE c.STATUS=2 AND c.award_time IS NOT NULL  AND c.award_time>=\'2022-07-01 00:00:00\' AND c.award_time<=\'2022-07-06 00:00:00\' AND c.proxy0_id=(SELECT id FROM m_account WHERE login_account=\'d0\') AND c.user_id IN(SELECT c.user_id AS \'账号\' FROM o_account_order AS c WHERE c.proxy0_id=(SELECT id FROM m_account WHERE login_account=\'d0\') AND c.STATUS=2 AND c.bet_type>1 AND c.award_time IS NOT NULL AND c.award_time>=\'2022-07-01 00:00:00\' AND c.award_time<=\'2022-07-06 00:00:00\' GROUP BY c.user_id,c.award_time ORDER BY c.award_time ASC) GROUP BY c.user_name,c.login_account,c.user_id,c.currency,sport_id,c.order_no,c.sport_id,c.company_retreat_proportion,c.level0_actual_percentage,c.level1_actual_percentage,c.level2_actual_percentage,c.level3_actual_percentage,c.level0_retreat_proportion,c.level1_retreat_proportion,c.level2_retreat_proportion,c.level3_retreat_proportion,c.award_time,c.efficient_amount,c.user_id,c.create_time ORDER BY c.create_time ASC"\n@f"SELECT away_team_name,bet_score,home_team_name,market_name,match_time,CASE is_live WHEN \'1\' THEN \'滚球盘\' WHEN \'3\' THEN \'早盘\' END \'盘口类型\',credit_odds,odds_type AS \'赔率类型\',order_no,outcome_name,specifier,tournament_name FROM o_account_order_match WHERE order_no in (SELECT c.order_no FROM o_account_order AS c LEFT JOIN u_user AS b ON b.id=c.user_id LEFT JOIN o_account_order_match AS a ON a.order_no=c.order_no WHERE c.STATUS=2 AND c.award_time IS NOT NULL  AND c.award_time>=\'{begin}\' AND c.award_time<=\'{end}\' AND c.proxy0_id=(SELECT id FROM m_account WHERE login_account=\'d0\') AND c.user_id IN(SELECT c.user_id AS \'账号\' FROM o_account_order AS c WHERE c.proxy0_id=(SELECT id FROM m_account WHERE login_account=\'d0\') AND c.STATUS=2 AND c.bet_type>1 AND c.award_time IS NOT NULL AND c.award_time>=\'2022-07-01 00:00:00\' AND c.award_time<=\'2022-07-06 00:00:00\' GROUP BY c.user_id,c.award_time ORDER BY c.award_time ASC) GROUP BY c.order_no) GROUP BY order_no,away_team_name,bet_score,home_team_name,market_name,match_time,is_live,credit_odds,odds_type,outcome_name,specifier,tournament_name,create_time ORDER BY create_time ASC"', 'account,orderNo', '测试通过\n2022-07-13 12:54:58']
        # sport_report_list =[{'account': 'd0d1d2d30q/fceshi016', 'betAmount': 40.0, 'betIp': '192.168.10.120', 'betIpAddress': '局域网', 'betMix': '3串4', 'betResult': '赢', 'betTime': '2022-06-27 08:43:09', 'betType': '复式串关', 'level0Commission': 0.0, 'level0CommissionRatio': 0.0, 'level0Percentage': 0.2, 'level0Total': -4.01, 'level0WinOrLose': -4.01, 'level1Commission': 0.0, 'level1CommissionRatio': 0.0, 'level1Percentage': 0.2, 'level1Total': -4.01, 'level1WinOrLose': -4.01, 'level2Commission': 0.0, 'level2CommissionRatio': 0.0, 'level2Percentage': 0.2, 'level2Total': -4.01, 'level2WinOrLose': -4.01, 'level3Commission': 0.0, 'level3CommissionRatio': 0.0, 'level3Percentage': 0.2, 'level3Total': -4.01, 'level3WinOrLose': -4.01, 'memberCommission': 0.0, 'memberCommissionRatio': 0.0, 'memberTotal': 20.09, 'memberWinOrLose': 20.09, 'mixNum': '3_4_0', 'name': '杜鑫test账号aq', 'odds': 22.68, 'oddsType': '2', 'options': [{'awayTeamName': '德利城', 'betScore': None, 'homeTeamName': '布咸美恩斯', 'marketName': '上半场 - 德利城 大/小', 'matchTime': '2022-07-01 14:45:00', 'matchType': '早盘', 'odds': 0.91, 'oddsType': '2', 'orderNo': 'XHwPqfvctxFR', 'outcomeName': '大0.5', 'specifier': 'total=0.5', 'tournamentName': '爱尔兰足球超级联赛'}, {'awayTeamName': '撒马尔罕迪纳摩', 'betScore': None, 'homeTeamName': '浩罕1912', 'marketName': '撒马尔罕迪纳摩 大/小', 'matchTime': '2022-06-28 11:00:00', 'matchType': '早盘', 'odds': 0.55, 'oddsType': '2', 'orderNo': 'XHwPqfvctxFR', 'outcomeName': '大0.5', 'specifier': 'total=0.5', 'tournamentName': '乌兹别克斯坦职业足球联赛'}, {'awayTeamName': '阿德莱得城', 'betScore': None, 'homeTeamName': '白城伍德维尔', 'marketName': '大/小', 'matchTime': '2022-07-02 01:30:00', 'matchType': '早盘', 'odds': 0.67, 'oddsType': '2', 'orderNo': 'XHwPqfvctxFR', 'outcomeName': '大2.5/3', 'specifier': 'total=2.75', 'tournamentName': '澳大利亚全国超级联赛,南澳大利亚'}, {'awayTeamName': 'Dunfermline Athletic FC', 'betScore': None, 'homeTeamName': 'Forfar Athletic FC', 'marketName': '大/小', 'matchTime': '2022-06-28 14:30:00', 'matchType': '早盘', 'odds': 1.03, 'oddsType': '2', 'orderNo': 'XHwPqfvctxFR', 'outcomeName': '大3/3.5', 'specifier': 'total=3.25', 'tournamentName': '国际俱乐部俱乐部友谊赛'}], 'orderNo': 'XHwPqfvctxFR', 'settlementTime': '2022-07-02 03:27:42', 'sportId': 'sr:sport:1', 'sportType': '足球', 'validAmount': 20.09, 'winOrLose': 20.09}]
        # sport_report_dict =['account', 'betAmount', 'betIp', 'betIpAddress', 'betMix', 'betResult', 'betTime', 'betType', 'level0Commission', 'level0CommissionRatio', 'level0Percentage', 'level0Total', 'level0WinOrLose', 'level1Commission', 'level1CommissionRatio', 'level1Percentage', 'level1Total', 'level1WinOrLose', 'level2Commission', 'level2CommissionRatio', 'level2Percentage', 'level2Total', 'level2WinOrLose', 'level3Commission', 'level3CommissionRatio', 'level3Percentage', 'level3Total', 'level3WinOrLose', 'memberCommission', 'memberCommissionRatio', 'memberTotal', 'memberWinOrLose', 'mixNum', 'name', 'odds', 'oddsType', 'options', 'orderNo', 'settlementTime', 'sportId', 'sportType', 'validAmount', 'winOrLose']
        # yyds=bc.sport_report_sql(begin="2022-07-01 00:00:00",end="2022-07-07 23:59:59", excel_report=excel_report, sport_report_dict=sport_report_dict,sport_report_list=sport_report_list)


        # A=[{'account': 'd0d1d2d3pe/fceshi0646', 'betAmount': 10.0, 'betIp': '192.168.10.120', 'betIpAddress': '局域网', 'betResult': '输', 'betTime': '2022-06-24 09:02:58', 'betType': '单注', 'level0Commission': 0.0, 'level0CommissionRatio': 0.0, 'level0Percentage': 0.2, 'level0Total': 2.0, 'level0WinOrLose': 2.0, 'level1Commission': 0.0, 'level1CommissionRatio': 0.0, 'level1Percentage': 0.2, 'level1Total': 2.0, 'level1WinOrLose': 2.0, 'level2Commission': 0.0, 'level2CommissionRatio': 0.0, 'level2Percentage': 0.2, 'level2Total': 2.0, 'level2WinOrLose': 2.0, 'level3Commission': 0.0, 'level3CommissionRatio': 0.0, 'level3Percentage': 0.2, 'level3Total': 2.0, 'level3WinOrLose': 2.0, 'memberCommission': 0.0, 'memberCommissionRatio': 0.0, 'memberTotal': -10.0, 'memberWinOrLose': -10.0, 'name': '杜鑫test账号pe', 'odds': 2.75, 'oddsType': '1', 'options': [{'awayTeamName': '海德堡联', 'betScore': None, 'homeTeamName': '阿东那', 'marketName': '独赢', 'matchTime': '2022-06-25 04:00:00', 'matchType': '早盘', 'odds': 2.75, 'oddsType': '1', 'orderNo': 'XH4AeukHvicj', 'outcomeName': '阿东那', 'specifier': '', 'tournamentName': '澳大利亚全国超级联赛,维多利亚'}], 'orderNo': 'XH4AeukHvicj', 'settlementTime': '2022-06-25 05:58:10', 'sportId': 'sr:sport:1', 'sportType': '足球', 'validAmount': 10.0, 'winOrLose': -10.0}]
        # B=[{'account': 'd0d1d2d39v/fceshi0247', 'betAmount': 10.0, 'betIp': '192.168.10.120', 'betIpAddress': '局域网', 'betResult': '输', 'betTime': '2022-06-24 08:40:18', 'betType': '单注', 'level0Commission': 0.0, 'level0CommissionRatio': 0.0, 'level0Percentage': 0.2, 'level0Total': 2.0, 'level0WinOrLose': 2.0, 'level1Commission': 0.0, 'level1CommissionRatio': 0.0, 'level1Percentage': 0.2, 'level1Total': 2.0, 'level1WinOrLose': 2.0, 'level2Commission': 0.0, 'level2CommissionRatio': 0.0, 'level2Percentage': 0.2, 'level2Total': 2.0, 'level2WinOrLose': 2.0, 'level3Commission': 0.0, 'level3CommissionRatio': 0.0, 'level3Percentage': 0.2, 'level3Total': 2.0, 'level3WinOrLose': 2.0, 'memberCommission': 0.0, 'memberCommissionRatio': 0.0, 'memberTotal': -10.0, 'memberWinOrLose': -10.0, 'name': '杜鑫test账号jv', 'odds': 2.28, 'oddsType': '1', 'options': [{'awayTeamName': '巴恩斯利', 'betScore': None, 'homeTeamName': '沃克索谱镇', 'marketName': '总入球', 'matchTime': '2022-06-25 10:00:00', 'matchType': '早盘', 'odds': 2.28, 'oddsType': '1', 'orderNo': 'XH4suBVmDm3E', 'outcomeName': '2-3', 'specifier': 'variant=sr:goal_range:7+', 'tournamentName': '国际俱乐部俱乐部友谊赛'}], 'orderNo': 'XH4suBVmDm3E', 'settlementTime': '2022-06-25 12:16:57', 'sportId': 'sr:sport:1', 'sportType': '足球', 'validAmount': 10.0, 'winOrLose': -10.0}]
        # yyds = bc.if_error(A=A,B=B)



