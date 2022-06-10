import time
from Crypto.Cipher import PKCS1_v1_5 as Cipher_pkcs1_v1_5
from Crypto.PublicKey import RSA
import requests
import json
import base64
import random
import datetime
from concurrent.futures import ThreadPoolExecutor
import openpyxl




#rsa加密账号、密码
def rsa_encrypt(data):
    '''
    rsa加密（encrypt）
    :param data:
    :return:
    '''
    msg = data.encode('utf-8')
    pub_key = "-----BEGIN PUBLIC KEY-----\nMFwwDQYJKoZIhvcNAQEBBQADSwAwSAJBAL1XuLmIZttk13hmAGVuXiKSfQggfVck" \
              "p+iNr9jBIxkmBBfmygJ9D5A7lhUbhBEY1SqyGNIHI1DsNLfxfRvW2EcCAwEAAQ==\n-----END PUBLIC KEY-----"
    rsa_key = RSA.importKey(pub_key)
    cipher = Cipher_pkcs1_v1_5.new(rsa_key)
    cipher_text = base64.b64encode(cipher.encrypt(msg)).decode('utf-8')
    # print(cipher_text)
    return cipher_text











def login(URL):
    #登0~登3登录获取token
    url = str(URL) + "/system/accountLogin"
    headers = {
        'Content-Type': 'application/json;charset=UTF-8'
    }
    # data = {"userName":rsa_encrypt(data=userName_list[i]),
    #         "password":rsa_encrypt(data=password_list[i]),
    #         "securityCode":securityCode_list[i],
    #         "loginDiv":555666}
    data = {"userName": rsa_encrypt(data=Dl_list[pkk]),
            "password": rsa_encrypt(data="Bfty123456"),
            "securityCode": "Bf123456",
            "loginDiv": 555666}
    session = requests.session()
    response = session.post(url=url, headers=headers, json=data)
    # 返回结果json转化
    results = json.loads(response.text)
    token = results['data']['token']
    id = results['data']['id']


    if results['code']==0:
       print("登录成功")

    # 过渡调用函数：
    url01 = str(URL) + "/system/userDetailByToken?"
    headers01 = {
        'Content-Type': 'application/json;charset=UTF-8',
        'LoginDiv': '555666',
        'Account_Login_Identify': token
    }
    data01 = {
        "token": token,
        "accountId": id
    }
    session = requests.session()
    response01 = session.get(url=url01, headers=headers01, params=data01)
    # 返回结果json转化
    results01 = json.loads(response01.text)
    # print(results01)

    sport_report(URL, token, list=sport_report_list)
    for i in sportId_list:
        sport_market_report(URL, token, list=sport_market_report_list, sportId=i)
    yyds_excel(ball_list=sport_market_report_title, list=sport_market_report_list, report=2, name='球类报表-详细')

    tournament_report(URL, token, list=tournament_report_list)
    match_report(URL, token, list=match_report_list)

    count=1
    for j in matchId_list:
        print(j,len(matchId_list),count)
        match_market_report(URL, token, list=match_market_report_list, matchId=j)
        count=count+1
    yyds_excel(ball_list=match_market_report_title, list=match_market_report_list, report=5, name='赛事报表-详细')
    multiterm_report(URL, token, list=multiterm_report_list)


def sport_report(URL, token,list):
    global sportId_list
    url = str(URL) + "/winOrLost/sport"
    headers = {
        'Content-Type': 'application/json;charset=UTF-8',
        'LoginDiv': '555666',
        'Account_Login_Identify': token
    }

    data = {"matchId":"",
            "sportId":"",
            "queryDateType":3,
            "begin":"2022-06-02",
            "end":"2022-06-08",
            "searchAccount":"",
            "page":1,"limit":50}
    session = requests.session()
    response = session.post(url=url, headers=headers, json=data)
    results = json.loads(response.text)
    yyds=results['data']['data']
    sportId_list=[]

    for i in range(0,len(yyds)):
        sportId=yyds[i]['sportId']
        sportId_list.append(sportId)
        print(sportId)
    print(sportId_list)
    for j in range(0,len(yyds)):
        list = []
        sportName = yyds[j]['sportName']
        list.append(sportName)
        allAmount = yyds[j]['allAmount']
        list.append(allAmount)
        allEfficient = yyds[j]['allEfficient']
        list.append(allEfficient)
        allBackwater = yyds[j]['allBackwater']
        list.append(allBackwater)
        memberWinLose = yyds[j]['memberWinLose']
        list.append(memberWinLose)
        memberBackwater = yyds[j]['memberBackwater']
        list.append(memberBackwater)
        memberFinal = yyds[j]['memberFinal']
        list.append(memberFinal)
        level3WinLose = yyds[j]['level3WinLose']
        list.append(level3WinLose)
        level3Backwater = yyds[j]['level3Backwater']
        list.append(level3Backwater)
        level3Final = yyds[j]['level3Final']
        list.append(level3Final)
        level2WinLose = yyds[j]['level2WinLose']
        list.append(level2WinLose)
        level2Backwater = yyds[j]['level2Backwater']
        list.append(level2Backwater)
        level2Final = yyds[j]['level2Final']
        list.append(level2Final)

        level1WinLose = yyds[j]['level1WinLose']
        list.append(level1WinLose)
        level1Backwater = yyds[j]['level1Backwater']
        list.append(level1Backwater)
        level1Final = yyds[j]['level1Final']

        list.append(level1Final)
        level0WinLose = yyds[j]['level0WinLose']
        list.append(level0WinLose)
        level0Backwater = yyds[j]['level0Backwater']
        list.append(level0Backwater)
        level0Final = yyds[j]['level0Final']
        list.append(level0Final)

        companyWinOrLose = yyds[j]['companyWinOrLose']
        list.append(companyWinOrLose)
        companyBackwaterAmount = yyds[j]['companyBackwaterAmount']
        list.append(companyBackwaterAmount)
        companyFinal = yyds[j]['companyFinal']
        list.append(companyFinal)


        sport_report_list.append(list)


    sport_report_title=['球类名称','总投注金额','总有效金额','总佣金','会员输赢','会员佣金','会员共计','三级代理输赢','三级代理佣金','三级代理共计','二级代理输赢','二级代理佣金','二级代理共计',
                        '一级代理输赢','一级代理佣金','一级代理共计','总代理输赢','总代理佣金','总代理共计',"公司输赢","公司佣金","公司共计"]
    print(f"sport_report_list:{sport_report_list}\n")

    yyds_excel(ball_list=sport_report_title, list=sport_report_list, report=1, name='球类报表')

    return sportId_list

def sport_market_report(URL, token, list,sportId):
    global sport_market_report_title
    url = str(URL) + "/winOrLost/market"
    headers = {
        'Content-Type': 'application/json;charset=UTF-8',
        'LoginDiv': '555666',
        'Account_Login_Identify': token
    }

    data = {"sportId":sportId,
            "queryDateType":3}
    session = requests.session()
    response = session.post(url=url, headers=headers, json=data)
    results = json.loads(response.text)
    yyds = results['data']['data']

    for j in range(0,len(yyds)):
        list = []
        sportName = yyds[j]['sportName']
        list.append(sportName)
        sportMarketName = yyds[j]['sportMarketName']
        list.append(sportMarketName)
        allAmount = yyds[j]['allAmount']
        list.append(allAmount)
        allEfficient = yyds[j]['allEfficient']
        list.append(allEfficient)
        allBackwater = yyds[j]['allBackwater']
        list.append(allBackwater)
        memberWinLose = yyds[j]['memberWinLose']
        list.append(memberWinLose)
        memberBackwater = yyds[j]['memberBackwater']
        list.append(memberBackwater)
        memberFinal = yyds[j]['memberFinal']
        list.append(memberFinal)
        level3WinLose = yyds[j]['level3WinLose']
        list.append(level3WinLose)
        level3Backwater = yyds[j]['level3Backwater']
        list.append(level3Backwater)
        level3Final = yyds[j]['level3Final']
        list.append(level3Final)
        level2WinLose = yyds[j]['level2WinLose']
        list.append(level2WinLose)
        level2Backwater = yyds[j]['level2Backwater']
        list.append(level2Backwater)
        level2Final = yyds[j]['level2Final']
        list.append(level2Final)

        level1WinLose = yyds[j]['level1WinLose']
        list.append(level1WinLose)
        level1Backwater = yyds[j]['level1Backwater']
        list.append(level1Backwater)
        level1Final = yyds[j]['level1Final']

        list.append(level1Final)
        level0WinLose = yyds[j]['level0WinLose']
        list.append(level0WinLose)
        level0Backwater = yyds[j]['level0Backwater']
        list.append(level0Backwater)
        level0Final = yyds[j]['level0Final']
        list.append(level0Final)

        companyWinOrLose = yyds[j]['companyWinOrLose']
        list.append(companyWinOrLose)
        companyBackwaterAmount = yyds[j]['companyBackwaterAmount']
        list.append(companyBackwaterAmount)
        companyFinal = yyds[j]['companyFinal']
        list.append(companyFinal)


        sport_market_report_list.append(list)

    sport_market_report_title = ['球类名称','盘口名称', '总投注金额', '总有效金额', '总佣金', '会员输赢', '会员佣金', '会员共计', '三级代理输赢', '三级代理佣金', '三级代理共计','二级代理输赢',
                          '二级代理佣金', '二级代理共计','一级代理输赢', '一级代理佣金', '一级代理共计', '总代理输赢', '总代理佣金', '总代理共计',"公司输赢","公司佣金","公司共计"]
    print(f"sport_market_report_list:{sport_market_report_list}\n")






def tournament_report(URL, token, list):
    url = str(URL) + "/winOrLost/tournament"
    headers = {
        'Content-Type': 'application/json;charset=UTF-8',
        'LoginDiv': '555666',
        'Account_Login_Identify': token
    }

    data = {"matchId":"",
            "sportId":"","queryDateType":3,
            "begin":"2022-06-02",
            "end":"2022-06-08",
            "searchAccount":"",
            "page":1,"limit":50}
    session = requests.session()
    response = session.post(url=url, headers=headers, json=data)
    results = json.loads(response.text)
    yyds = results['data']['data']

    for j in range(0,len(yyds)):
        list = []
        sportName = yyds[j]['sportName']
        list.append(sportName)
        tournamentName = yyds[j]['tournamentName']
        list.append(tournamentName)
        allAmount = yyds[j]['allAmount']
        list.append(allAmount)
        allEfficient = yyds[j]['allEfficient']
        list.append(allEfficient)
        allBackwater = yyds[j]['allBackwater']
        list.append(allBackwater)
        memberWinLose = yyds[j]['memberWinLose']
        list.append(memberWinLose)
        memberBackwater = yyds[j]['memberBackwater']
        list.append(memberBackwater)
        memberFinal = yyds[j]['memberFinal']
        list.append(memberFinal)
        level3WinLose = yyds[j]['level3WinLose']
        list.append(level3WinLose)
        level3Backwater = yyds[j]['level3Backwater']
        list.append(level3Backwater)
        level3Final = yyds[j]['level3Final']
        list.append(level3Final)
        level2WinLose = yyds[j]['level2WinLose']
        list.append(level2WinLose)
        level2Backwater = yyds[j]['level2Backwater']
        list.append(level2Backwater)
        level2Final = yyds[j]['level2Final']
        list.append(level2Final)

        level1WinLose = yyds[j]['level1WinLose']
        list.append(level1WinLose)
        level1Backwater = yyds[j]['level1Backwater']
        list.append(level1Backwater)
        level1Final = yyds[j]['level1Final']

        list.append(level1Final)
        level0WinLose = yyds[j]['level0WinLose']
        list.append(level0WinLose)
        level0Backwater = yyds[j]['level0Backwater']
        list.append(level0Backwater)
        level0Final = yyds[j]['level0Final']
        list.append(level0Final)

        companyWinOrLose = yyds[j]['companyWinOrLose']
        list.append(companyWinOrLose)
        companyBackwaterAmount = yyds[j]['companyBackwaterAmount']
        list.append(companyBackwaterAmount)
        companyFinal = yyds[j]['companyFinal']
        list.append(companyFinal)
        tournament_report_list.append(list)

    tournament_report_title = ['球类名称','联赛名称', '总投注金额', '总有效金额', '总佣金', '会员输赢', '会员佣金', '会员共计', '三级代理输赢', '三级代理佣金',
                                 '三级代理共计','二级代理输赢',
                                 '二级代理佣金', '二级代理共计',
                                 '一级代理输赢', '一级代理佣金', '一级代理共计', '总代理输赢', '总代理佣金', '总代理共计','公司输赢', '公司佣金', '公司共计']
    print(f"tournament_report_list:{tournament_report_list}\n")
    yyds_excel(ball_list=tournament_report_title, list=tournament_report_list, report=3, name='联赛报表')





def match_report(URL, token, list):
    global matchId_list
    url = str(URL) + "/winOrLost/match"
    headers = {
        'Content-Type': 'application/json;charset=UTF-8',
        'LoginDiv': '555666',
        'Account_Login_Identify': token
    }

    data = {"matchId":"",
            "sportId":"",
            "queryDateType":3,
            "begin":"2022-06-02",
            "end":"2022-06-08",
            "searchAccount":"",
            "page":1,"limit":50}
    session = requests.session()
    response = session.post(url=url, headers=headers, json=data)
    results = json.loads(response.text)
    yyds = results['data']['data']
    matchId_list = []

    for i in results['data']['data']:
        matchId = i['matchId']
        matchId_list.append(matchId)

    for j in range(0,len(yyds)):
        list = []
        sportName = yyds[j]['sportName']
        list.append(sportName)
        awayTeamName = yyds[j]['awayTeamName']
        homeTeamName = yyds[j]['homeTeamName']
        list.append((str(awayTeamName)+"VS"+str(homeTeamName)))
        allAmount = yyds[j]['allAmount']
        list.append(allAmount)
        allEfficient = yyds[j]['allEfficient']
        list.append(allEfficient)
        allBackwater = yyds[j]['allBackwater']
        list.append(allBackwater)
        memberWinLose = yyds[j]['memberWinLose']
        list.append(memberWinLose)
        memberBackwater = yyds[j]['memberBackwater']
        list.append(memberBackwater)
        memberFinal = yyds[j]['memberFinal']
        list.append(memberFinal)
        level3WinLose = yyds[j]['level3WinLose']
        list.append(level3WinLose)
        level3Backwater = yyds[j]['level3Backwater']
        list.append(level3Backwater)
        level3Final = yyds[j]['level3Final']
        list.append(level3Final)
        level2WinLose = yyds[j]['level2WinLose']
        list.append(level2WinLose)
        level2Backwater = yyds[j]['level2Backwater']
        list.append(level2Backwater)
        level2Final = yyds[j]['level2Final']
        list.append(level2Final)

        level1WinLose = yyds[j]['level1WinLose']
        list.append(level1WinLose)
        level1Backwater = yyds[j]['level1Backwater']
        list.append(level1Backwater)
        level1Final = yyds[j]['level1Final']

        list.append(level1Final)
        level0WinLose = yyds[j]['level0WinLose']
        list.append(level0WinLose)
        level0Backwater = yyds[j]['level0Backwater']
        list.append(level0Backwater)
        level0Final = yyds[j]['level0Final']
        list.append(level0Final)

        companyWinOrLose = yyds[j]['companyWinOrLose']
        list.append(companyWinOrLose)
        companyBackwaterAmount = yyds[j]['companyBackwaterAmount']
        list.append(companyBackwaterAmount)
        companyFinal = yyds[j]['companyFinal']
        list.append(companyFinal)
        match_report_list.append(list)



    match_report_title = ['球类名称','赛事名称', '总投注金额', '总有效金额', '总佣金', '会员输赢', '会员佣金', '会员共计', '三级代理输赢', '三级代理佣金',
                               '三级代理共计','二级代理输赢',
                               '二级代理佣金', '二级代理共计',
                               '一级代理输赢', '一级代理佣金', '一级代理共计', '总代理输赢', '总代理佣金', '总代理共计', '公司输赢', '公司佣金', '公司共计']
    print(f"match_report_list:{match_report_list}\n")
    yyds_excel(ball_list=match_report_title, list=match_report_list, report=4, name='赛事报表')

    return matchId_list

def match_market_report(URL, token, list,matchId):
    global match_market_report_title
    url = str(URL) + "/winOrLost/market"
    headers = {
        'Content-Type': 'application/json;charset=UTF-8',
        'LoginDiv': '555666',
        'Account_Login_Identify': token
    }

    data = {"matchId":matchId,
            "queryDateType":3}
    session = requests.session()
    response = session.post(url=url, headers=headers, json=data)
    results = json.loads(response.text)
    yyds = results['data']['data']

    for j in range(0,len(yyds)):
        list = []
        sportName = yyds[j]['sportName']
        list.append(sportName)
        sportMarketName = yyds[j]['sportMarketName']
        list.append(sportMarketName)
        allAmount = yyds[j]['allAmount']
        list.append(allAmount)
        allEfficient = yyds[j]['allEfficient']
        list.append(allEfficient)
        allBackwater = yyds[j]['allBackwater']
        list.append(allBackwater)
        memberWinLose = yyds[j]['memberWinLose']
        list.append(memberWinLose)
        memberBackwater = yyds[j]['memberBackwater']
        list.append(memberBackwater)
        memberFinal = yyds[j]['memberFinal']
        list.append(memberFinal)
        level3WinLose = yyds[j]['level3WinLose']
        list.append(level3WinLose)
        level3Backwater = yyds[j]['level3Backwater']
        list.append(level3Backwater)
        level3Final = yyds[j]['level3Final']
        list.append(level3Final)
        level2WinLose = yyds[j]['level2WinLose']
        list.append(level2WinLose)
        level2Backwater = yyds[j]['level2Backwater']
        list.append(level2Backwater)
        level2Final = yyds[j]['level2Final']
        list.append(level2Final)

        level1WinLose = yyds[j]['level1WinLose']
        list.append(level1WinLose)
        level1Backwater = yyds[j]['level1Backwater']
        list.append(level1Backwater)
        level1Final = yyds[j]['level1Final']

        list.append(level1Final)
        level0WinLose = yyds[j]['level0WinLose']
        list.append(level0WinLose)
        level0Backwater = yyds[j]['level0Backwater']
        list.append(level0Backwater)
        level0Final = yyds[j]['level0Final']
        list.append(level0Final)

        companyWinOrLose = yyds[j]['companyWinOrLose']
        list.append(companyWinOrLose)
        companyBackwaterAmount = yyds[j]['companyBackwaterAmount']
        list.append(companyBackwaterAmount)
        companyFinal = yyds[j]['companyFinal']
        list.append(companyFinal)
        match_market_report_list.append(list)

    match_market_report_title = ['球类名称','盘口名称', '总投注金额', '总有效金额', '总佣金', '会员输赢', '会员佣金', '会员共计', '三级代理输赢', '三级代理佣金',
                              '三级代理共计','二级代理输赢',
                              '二级代理佣金', '二级代理共计',
                              '一级代理输赢', '一级代理佣金', '一级代理共计', '总代理输赢', '总代理佣金', '总代理共计','公司输赢', '公司佣金', '公司共计']



def multiterm_report(URL,token,list):

    url = str(URL) + "/winOrLost/multiterm"
    headers = {
        'Content-Type': 'application/json;charset=UTF-8',
        'LoginDiv': '555666',
        'Account_Login_Identify': token
    }

    data = {"matchId":"",
            "sportId":"",
            "queryDateType":3,
            "begin":"2022-06-02",
            "end":"2022-06-08",
            "searchAccount":"",
            "page":1,
            "limit":50}
    session = requests.session()
    response = session.post(url=url, headers=headers, json=data)
    results = json.loads(response.text)
    yyds = results['data']['data']

    for j in range(0,len(yyds)):
        list = []
        userName = yyds[j]['userName']
        loginAccount = yyds[j]['loginAccount']
        list.append((str(userName)+"/")+str(loginAccount))
        allAmount = yyds[j]['allAmount']
        list.append(allAmount)
        allEfficient = yyds[j]['allEfficient']
        list.append(allEfficient)
        allBackwater = yyds[j]['allBackwater']
        list.append(allBackwater)
        memberWinLose = yyds[j]['memberWinLose']
        list.append(memberWinLose)
        memberBackwater = yyds[j]['memberBackwater']
        list.append(memberBackwater)
        memberFinal = yyds[j]['memberFinal']
        list.append(memberFinal)
        level3WinLose = yyds[j]['level3WinLose']
        list.append(level3WinLose)
        level3Backwater = yyds[j]['level3Backwater']
        list.append(level3Backwater)
        level3Final = yyds[j]['level3Final']
        list.append(level3Final)
        level2WinLose = yyds[j]['level2WinLose']
        list.append(level2WinLose)
        level2Backwater = yyds[j]['level2Backwater']
        list.append(level2Backwater)
        level2Final = yyds[j]['level2Final']
        list.append(level2Final)

        level1WinLose = yyds[j]['level1WinLose']
        list.append(level1WinLose)
        level1Backwater = yyds[j]['level1Backwater']
        list.append(level1Backwater)
        level1Final = yyds[j]['level1Final']

        list.append(level1Final)
        level0WinLose = yyds[j]['level0WinLose']
        list.append(level0WinLose)
        level0Backwater = yyds[j]['level0Backwater']
        list.append(level0Backwater)
        level0Final = yyds[j]['level0Final']
        list.append(level0Final)

        companyWinOrLose = yyds[j]['companyWinOrLose']
        list.append(companyWinOrLose)
        companyBackwaterAmount = yyds[j]['companyBackwaterAmount']
        list.append(companyBackwaterAmount)
        companyFinal = yyds[j]['companyFinal']
        list.append(companyFinal)
        multiterm_report_list.append(list)

    multiterm_report_title = ['账号/登入账号', '总投注金额', '总有效金额', '总佣金', '会员输赢', '会员佣金', '会员共计', '三级代理输赢', '三级代理佣金',
                          '三级代理共计','二级代理输赢',
                          '二级代理佣金', '二级代理共计',
                          '一级代理输赢', '一级代理佣金', '一级代理共计', '总代理输赢', '总代理佣金', '总代理共计', '公司输赢', '公司佣金', '公司共计']
    print(f"multiterm_report_list:{multiterm_report_list}\n")
    yyds_excel(ball_list=multiterm_report_title, list=multiterm_report_list, report=6, name='混合串关')



def yyds_excel(ball_list,list,report,name):
    # 获取 工作簿对象
    workbook = openpyxl.load_workbook(excel[pkk])
    print("读取表格成功")
    # shenames = workbook.get_sheet_names()
    shenames = workbook.sheetnames
    print("整个表格所有表名:", shenames)  # ['各省市', '测试表']

    if len(shenames) > 10:
        pass
    else:
        for i in range(len(shenames), 11):
            workbook.create_sheet()

    worksheet = workbook[shenames[int(report)-1]]
    worksheet.title = name

    # 获取该表相应的行数和列数
    rows = worksheet.max_row
    columns = worksheet.max_column
    print("所在表组成有：" + str(rows) + "行", str(columns) + "列")

    # ball_list = []
    # list = []



    content_A1 = worksheet.cell(1, 1).value
    if content_A1 == None:
        for i in range(1, len(ball_list)+1):
            worksheet.cell(1, i, ball_list[i - 1])
        for j in range(2, len(list)+2):
                for k in range(1, len(ball_list)+1):
                    if worksheet.cell(j, k).value == None:
                        ppds=(list[j - 2][k - 1])
                        worksheet.cell(j, k, ppds)
                    else:
                        pass
        # for pp in range(0,999):
        #     if worksheet.cell(pp, 1).value == None:
        #         worksheet.cell(pp, 1, '合计')
        #         for kk in range(2, len(list)+2):
        #             worksheet.cell(pp, kk, )

        print(f"用户{Dl_list[pkk]}的{name}表写入完成")
    else:
        rows = worksheet.max_row
        columns = worksheet.max_column
        for i in range(1, int(rows) + 1):
            for j in range(1, int(columns) + 1):
                worksheet.cell(i, j).value = ""
        print(f"用户{Dl_list[pkk]}的{name}表已有内容无法写入,已清空")

    workbook.save(filename=save_excel[pkk])
    workbook.close()


def excel_clear():
    # 获取 工作簿对象
    workbook = openpyxl.load_workbook(excel[pkk])
    print("读取表格成功")
    # shenames = workbook.get_sheet_names()
    shenames = workbook.sheetnames
    print("整个表格所有表名:", shenames)  # ['各省市', '测试表']

    if len(shenames) > 10:
        pass
    else:
        for i in range(len(shenames), 11):
            workbook.create_sheet()
    for i in range(0,6):
        worksheet = workbook[shenames[i]]
        # worksheet.title = name

        # 获取该表相应的行数和列数
        rows = worksheet.max_row
        columns = worksheet.max_column
        print("所在表组成有：" + str(rows) + "行", str(columns) + "列")

        # ball_list = []
        # list = []



        content_A1 = worksheet.cell(1, 1).value
        if content_A1 == None:
            pass
        else:
            for i in range(1, rows+1):
                for j in range(1, columns+1):
                    worksheet.cell(i, j).value = ""
            workbook.save(filename=save_excel[pkk])
            workbook.close()
    print(f"{save_excel[pkk]}   已清空")
















if __name__=='__main__':
    # MDE环境
    URL="https://mdesearch.betf.best"

    # multiterm_report_list = []
    # match_market_report_list = []
    # match_report_list = []
    # tournament_report_list = []
    # sport_market_report_list = []
    # sport_report_list = []


    for pkk in range(0,4):
        multiterm_report_list = []
        match_market_report_list = []
        match_report_list = []
        tournament_report_list = []
        sport_market_report_list = []
        sport_report_list = []
        Dl_list=["d0","d10","d2","d3"]
        excel=[r"C:\test\登0_report.xlsx",r"C:\test\登1_report.xlsx",r"C:\test\登2_report.xlsx",r"C:\test\登3_report.xlsx"]
        save_excel = ["C:\\test\\登0_report.xlsx", "C:\\test\\登1_report.xlsx", "C:\\test\\登2_report.xlsx","C:\\test\\登3_report.xlsx"]
        login(URL)

        # excel_clear()

