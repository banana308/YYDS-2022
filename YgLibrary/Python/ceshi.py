#!/usr/bin/python
# -*- coding: utf-8 -*-
import os, sys
import datetime
import time
import openpyxl,xlrd
from os.path import  join,abspath,dirname

from openpyxl import load_workbook
from openpyxl import workbook
import xlsxwriter


# list_A=['8','7','101','10','9','11','100']

# list_y=[]
# list01=[]
# #写入text
# path = "C:\\test\\michid.txt"
# f = open(path, mode="r")
# lines=f.readlines()
# for lines in lines:
#     list_y.append(lines.strip('\n'))
#     for k in range(0,len(list_y)):
#         if  list_y[k] in list01:
#             pass
#         else:
#             list01.append(list_y[k])
# print(len(list01))
# print(list01)


# for i in range(0,len(order_no02)):
#     if order_no02[i] in order_no01:
#         order_no01.remove(order_no02[i])
#     else:
#         pass
# print(len(order_no01))
# print(order_no01)
# ['XgPLKTrmDbMh', 'XgPLMKw4QUg5', 'XgPLNg4MK3Gu', 'XgPLNgsJtu4V', 'XgPLPetGNzSP']



# list=[]
# for i in range(1,1001):
#     name = "fceshi0" + str(i)
#     # name="sr:match:27885276"
#     # name="6210"
#     # path01 = "C:\\test\\test_jmeter\\matchId.txt"
#     path01 = "C:\\test\正式版\\user.txt"
#     f = open(path01, 'a')
#     f.write( name+ "\n")
#     list.append(name)
#     if i==1000:
#         print(len(list))
#     else:
#         pass



# lsit = [2, 3, 4, 5, 6, 7, 8, 9]
# for i in range(0, len(lsit) - 1):
#     a = sum(lsit[0:i])
#     b = sum(lsit[0:(i + 1)])
#
#     print(a, b, " ", str(lsit[i]), "串1")



# def get_time(number):
#     for number in range(number, 0, -1):
#         time.sleep(1)
#         print(number,"秒",end=" end")
#
#
# get_time(15)


# all_user_old=[{},{},{},{}]
#
# userName_list=['yy01','yy02','yy03','yy04',]
# psd_list=['yy01','yy02','yy03','yy04',]
#
# for i in range(0,4):
#     all_user_old[i][f'登{i}的账号']=userName_list[i]
#     all_user_old[i][f'登{i}的密码'] = psd_list[i]
#     print(all_user_old[i])
#
# print(all_user_old)











# list=['XfL2iMdzsph8', 'XfL2j2Ydpbwe', 'XfL2jgUUqwdX', 'XfL2jvN4GaeV', 'XfL2jKeJPJjD', 'XfL2jWW4g645', 'XfL2keq5S3B9', 'XfL2ks57jXAQ', 'XfL2kF9c7zQd', 'XfL2mCpAFQjE', 'XfL2mRsS6tWp', 'XfL2n6x9Mewd', 'XfL2nhZi6J5F', 'XfL2nxtUhLxV', 'XfL2nKrYQmr7', 'XfL2nVB26kAz', 'XfL2p8YMWThh', 'XfL2pidqLuZ8', 'XfL2pvRaSGpZ', 'XfL2pHJbakrY', 'XfL2pUJSnP2i', 'XfL2qaJbZM8Z', 'XfL2qkhdkGWB', 'XfL2qybJeSYL', 'XfL2qL48Y99D', 'XfL2qWqKbeqh', 'XfL2r7rjY5m7', 'XfL2rfAiNatf', 'XfL2rsmacZ5c', 'XfL2rCTXVB8H', 'XfL2rN7KjPzS', 'XfL2sMVs2BhQ', 'XfL2tkujJXcM', 'XfL2tCkG42Ge', 'XfL2tSitdfBi', 'XfL2u3MATvjS', 'XfL2ucMntxUv', 'XfL2umfnyBug', 'XfL2uwDrmgnr', 'XfL2uEYSZs37', 'XfL2uX86cw7Q', 'XfL2v7MxrjBG', 'XfL2vq9dZubz', 'XfL2vBJBnMYz', 'XfL2vNryBAjz', 'XfL2w22Wm8pV', 'XfL2wjap4UAz', 'XfL2wwijcP2J', 'XfL2wLvtVjSq', 'XfL2wVjPGNGc', 'XfL2x7ySJjaK', 'XfL2xjkvWFbY', 'XfL2xuk4wvaq', 'XfL2xGSUwPmj', 'XfL2xXiuw2k7', 'XfL2ybTvQxvu', 'XfL2yqaPjGrD', 'XfL2yBZyfn4a', 'XfL2yP9tb7Q7', 'XfL2yYqTPNKc', 'XfL2z99GA6Bf', 'XfL2zhSAjcUM', 'XfL2zt4evFkD', 'XfL2zDFNcjwY', 'XfL2zNvYjyn6', 'XfL2A2aMQGKq', 'XfL2Ad3HHiQC', 'XfL2AqxSurL7', 'XfL2ABKzemCT', 'XfL2AQMKhNga', 'XfL2B4QqZLpG', 'XfL2BgCRWCvD', 'XfL2Bup7rpzi', 'XfL2BJatHLgF', 'XfL2BVYn7mGY', 'XfL2C8TazyPr', 'XfL2CjJbY88d', 'XfL2CxhnNxHL', 'XfL2CKbiVNSg', 'XfL2CVDDJVk3', 'XfL2D7tAU6wM', 'XfL2Dj5LbtCP', 'XfL2DwpRhchK', 'XfL2DGvhfBE3', 'XfL2DSpZaJgU', 'XfL2E53qTkML', 'XfL2EgT9sWkL', 'XfL2EuJ5t66Z', 'XfL2EFncpqRc', 'XfL2ESztAzxL', 'XfL2F4yK37YG', 'XfL2FefEfSej', 'XfL2FpneCV9r', 'XfL2FxsM7eVa', 'XfL2FFW5a2dE', 'XfL2FQRfgxX5', 'XfL2FZnmFqfs', 'XfL2G9KzGiLn', 'XfL2GhddR2Zu', 'XfL2GrmB6NrU']
# lsit01=['XfL2iMdzsph8', 'XfL2j2Ydpbwe', 'XfL2jgUUqwdX', 'XfL2jvN4GaeV', 'XfL2jKeJPJjD', 'XfL2jWW4g645', 'XfL2keq5S3B9', 'XfL2ks57jXAQ', 'XfL2kF9c7zQd', 'XfL2mCpAFQjE', 'XfL2mRsS6tWp', 'XfL2n6x9Mewd', 'XfL2nhZi6J5F', 'XfL2nxtUhLxV', 'XfL2nKrYQmr7', 'XfL2nVB26kAz', 'XfL2p8YMWThh', 'XfL2pidqLuZ8', 'XfL2pvRaSGpZ', 'XfL2pHJbakrY', 'XfL2pUJSnP2i', 'XfL2qaJbZM8Z', 'XfL2qkhdkGWB', 'XfL2qybJeSYL', 'XfL2qL48Y99D', 'XfL2qWqKbeqh', 'XfL2r7rjY5m7', 'XfL2rfAiNatf', 'XfL2rsmacZ5c']
#
# for i in range(0,len(lsit01)):
#     if lsit01[i] in list:
#         list.remove(lsit01[i])
#     else:
#         pass
# print(len(list))
# print(list)





# list=["1","2","3","4","5","6",]




from concurrent.futures import ThreadPoolExecutor
import time

# def spider(page):
#     time.sleep(page)
#     print(f"crawl task{page} finished")
#     return page
#
# with ThreadPoolExecutor(max_workers=5) as t:  # 创建一个最大容纳数量为5的线程池
#         task1 = t.submit(spider, 1)
#         task2 = t.submit(spider, 2)  # 通过submit提交执行的函数到线程池中
#         task3 = t.submit(spider, 3)
#
#         print(f"task1: {task1.done()}")  # 通过done来判断线程是否完成
#         print(f"task2: {task2.done()}")
#         print(f"task3: {task3.done()}")
#
#         time.sleep(5)
#         print(f"task1: {task1.done()}")
#         print(f"task2: {task2.done()}")
#         print(f"task3: {task3.done()}")
#         print(task1.result())  # 通过result来获取返回值

# # num_lsit = ["复式3串4": [2_3, 3_4],"复式4串11": {[2_6, 3_4, 4_11]},"复式5串26": {[2_10, 3_10, 4_5, 5_26]},"复式6串57": {[2_15, 3_20, 4_15,5_6,6_57]}]
num_lsit = [{"复式3串4":["2_3_0","3_4_1"]},{"复式4串11":["2_6_0","3_4_0","4_11_1"]},{"复式5串25": ["2_10_0", "3_10_0", "4_5_0", "5_26_1"]},{"复式6串57": ["2_15_0", "3_20_0", "4_15_0","5_6_0","6_57_1"]}]




# new_list=[]
# old_list=[]
# tuple=(('足球', '阿尔及利亚甲级联赛', datetime.datetime(2022, 6, 8, 13, 45), '早盘   阿尔及利亚甲级联赛   ES Setif(主队) VS 巴拉特(客队)', '独赢', 0, 0, 0, 0, 1, 0), ('足球', '阿尔及利亚甲级联赛', datetime.datetime(2022, 6, 8, 13, 45), '早盘   阿尔及利亚甲级联赛   ES Setif(主队) VS 巴拉特(客队)', '独赢', 0, 0, 0, 0, 0, 1), ('足球', '阿尔及利亚甲级联赛', datetime.datetime(2022, 6, 8, 13, 45), '早盘   阿尔及利亚甲级联赛   ES Setif(主队) VS 巴拉特(客队)', '独赢', 0, 0, 0, 0, 0, 0))
# for i in range(0,len(tuple)):
#     old_list.append(list(tuple[i]))
#     print(tuple[i])
#     if i==0:
#         new_list.append(list(old_list[i]))
#     else:
#         for j in range(0,len(new_list)):
#             old = str(old_list[i][0]) + str(old_list[i][1]) + str(old_list[i][2]) + str(old_list[i][3]) + str(old_list[i][4])
#             new=str(new_list[j][0]) + str(new_list[j][1]) + str(new_list[j][2]) + str(new_list[j][3]) + str(new_list[j][4])
#             if old==new:
#                 new_list[j][5]=int(old_list[i][5])+ int(new_list[j][5])
#                 new_list[j][6]=int(old_list[i][6])+ int(new_list[j][6])
#                 new_list[j][7]=int(old_list[i][7])+ int(new_list[j][7])
#                 new_list[j][8]=int(old_list[i][8])+ int(new_list[j][8])
#                 new_list[j][9]=int(old_list[i][9])+ int(new_list[j][9])
#                 new_list[j][10]=int(old_list[i][10])+ int(new_list[j][10])
#             else:
#                 new_list.append(list(old_list[i]))
#
# for kk in new_list:
#     print(kk)





#
#
# # 获取 工作簿对象
# workbook = openpyxl.load_workbook(r"C:\test\test_ball.xlsx")
# print("读取表格成功")
# # shenames = workbook.get_sheet_names()
# shenames=workbook.sheetnames
# print("整个表格所有表名:",shenames)  #['各省市', '测试表']
#
# if len(shenames)>10:
#     pass
# else:
#     for i in range(len(shenames),11):
#         workbook.create_sheet()
#
# worksheet=workbook[shenames[2]]
# worksheet.title = "yyy"
#
#
#
# active=workbook.active
# active.merged_cells(sheet=active,range_string='A1：B3',start_row=5,start_column=5,end_row=5,end_column=5)
# worksheet.cell['A1:B3']="一级代理"
#
#
#
#
#
# #获取该表相应的行数和列数
# rows = worksheet.max_row
# columns = worksheet.max_column
# print("所在表组成有："+str(rows)+"行",str(columns)+"列")
#
#
# ball_list=['#','球类','开赛时间','赛事','0让球盘/主队','0让球盘/客队','0大小盘/大盘','0大小盘/小盘','0独赢/1','0独赢/X','0独赢/2','1让球盘/主队','1让球盘/客队','1大小盘/大盘','1大小盘/小盘','1独赢/1','1独赢/X','1独赢/2']
# list=[['足球', datetime.datetime(2022, 6, 2, 19, 30), '早盘   美国美足联乙级联赛   南佐治亚州风暴二队(主队) VS 桃树市MOBA(客队)', '独赢', 0, 0, 0, 0, 3, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 13, 0), '早盘   土耳其甲级联赛   Bandirmaspor(主队) VS Istanbulspor AS(客队)', '上半场 - 让球', 0, 0, 0, 0, 0, 0, 0, 2, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 10, 30), '早盘   伊朗足球甲级联赛   Rayka Babol FC(主队) VS 赛帕(客队)', '大/小', 0, 0, 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 17, 0), '早盘   哥伦比亚乙级联赛   Real Cartagena FC(主队) VS Itagui Leones FC(客队)', '让球', 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 7, 9, 15, 0), '早盘   国际欧洲锦标赛，女子   荷兰(主队) VS 瑞典(客队)', '上半场 - 独赢', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0], ['足球', datetime.datetime(2022, 6, 11, 3, 0), '早盘   澳大利亚全国超级联赛,西澳大利亚   Stirling Macedonia FC(主队) VS 阿马达尔(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 28, 13, 0), '早盘   瑞典超甲级联赛   Halmstads BK(主队) VS Trelleborgs FF(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 11, 3, 0), '早盘   澳大利亚全国超级联赛,西澳大利亚   湾水市(主队) VS 珀斯光荣(青年)(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 11, 0), '早盘   国际青年U23 AFC Asian Cup   泰国(主队) VS 越南(客队)', '上半场 - 让球', 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 17, 0), '早盘   哥伦比亚乙级联赛   Real Cartagena FC(主队) VS Itagui Leones FC(客队)', '上半场 - 独赢', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 11, 0), '早盘   突尼斯杯   比瑟汀(主队) VS 突尼斯施塔德(客队)', '大/小', 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 16, 0), '早盘   国际中北美洲及加勒比海国家联赛   安圭拉(主队) VS 多米尼加(客队)', '上半场 - 让球', 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 11, 30), '早盘   伊朗职业联赛   阿巴丹石油工业(主队) VS Havadar SC(客队)', '大/小', 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 9, 0), '早盘   国际青年U23 AFC Asian Cup   韩国(主队) VS 马来西亚(客队)', '大/小', 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 8, 10, 15, 0), '早盘   国际俱乐部欧洲超级杯   皇家马德里(主队) VS Eintracht Frankfurt(客队)', '上半场 - 让球', 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 7, 8, 15, 0), '早盘   国际欧洲锦标赛，女子   德国(主队) VS 丹麦(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 7, 18, 0), '早盘   巴西乙级联赛   GO维拉诺瓦(主队) VS SC布吕斯克(客队)', '上半场 - 独赢', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0], ['足球', datetime.datetime(2022, 6, 26, 9, 0), '早盘   瑞典超级联赛   Mjallby AIF(主队) VS IFK Norrkoping FK(客队)', '上半场 - 大/小', 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 26, 9, 0), '早盘   瑞典超甲级联赛   Dalkurd FF(主队) VS AFC Eskilstuna(客队)', '上半场 - 大/小', 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 11, 25, 15, 0), '早盘   国际世界杯   英格兰(主队) VS 美国(客队)', '大/小', 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 19, 30), '早盘   委内瑞拉甲级联赛   葡萄牙人(主队) VS 拉瓜伊拉(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 10, 30), '早盘   伊朗足球甲级联赛   克尔曼(主队) VS 贾姆帕斯祖洛比(客队)', '大/小', 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 7, 7, 15, 0), '早盘   国际欧洲锦标赛，女子   挪威(主队) VS 北爱尔兰(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 18, 6, 0), '早盘   日本J联赛   广岛三箭(主队) VS Cerezo Osaka(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 11, 22, 12, 0), '早盘   国际世界杯   墨西哥(主队) VS Poland(客队)', '上半场 - 让球', 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 15, 0), '早盘   国际非洲国家杯资格赛   突尼斯(主队) VS 赤道几内亚(客队)', '上半场 - 大/小', 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 8, 14, 45), '早盘   欧足协国际联赛   比利时(主队) VS Poland(客队)', '上半场 - 独赢', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 13, 0), '早盘   冰岛足球超级联赛，女子   雷克雅未克瓦鲁尔(主队) VS 韦斯特曼纳群岛(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 13, 0), '早盘   丹麦足球丙级联赛，第4池   克杰勒鲁普IF(主队) VS 富勒巴肯KFUM(客队)', '大/小', 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 10, 11, 45), '早盘   阿尔及利亚甲级联赛   NA侯赛因德尔(主队) VS 比斯克拉(客队)', '独赢', 0, 0, 0, 0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 10, 11, 45), '早盘   阿尔及利亚甲级联赛   NA侯赛因德尔(主队) VS 比斯克拉(客队)', '上半场 - 独赢', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1], ['足球', datetime.datetime(2022, 6, 28, 13, 0), '早盘   瑞典超甲级联赛   Norrby IF(主队) VS Ostersunds FK(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 7, 21, 0), '早盘   国际中北美洲及加勒比海国家联赛   牙买加(主队) VS Suriname(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 11, 23, 9, 0), '早盘   国际世界杯   德国(主队) VS 日本(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 17, 6, 0), '早盘   韩国K1联赛   金泉尚武足球俱乐部(主队) VS 水原(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 7, 12, 9, 0), '早盘   国际俱乐部俱乐部友谊赛   Manchester United(主队) VS Liverpool FC(客队)', '上半场 - 大/小', 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 11, 21, 9, 0), '早盘   国际世界杯   英格兰(主队) VS 伊朗(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 10, 30), '早盘   伊朗足球甲级联赛   Rayka Babol FC(主队) VS 赛帕(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 29, 14, 0), '早盘   国际俱乐部解放者杯   托利马(主队) VS 法林明高RJ(客队)', '上半场 - 独赢', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 13, 30), '早盘   国际青年U21欧洲锦标赛预选赛   卢森堡(主队) VS 瑞典(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 10, 14, 45), '早盘   爱尔兰甲级联赛   亚隆城(主队) VS Cobh Ramblers(客队)', '上半场 - 独赢', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0], ['足球', datetime.datetime(2022, 6, 27, 13, 0), '早盘   瑞典超甲级联赛   兰斯干拿(主队) VS Orebro SK(客队)', '大/小', 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 26, 9, 0), '早盘   瑞典超级联赛   Mjallby AIF(主队) VS IFK Norrkoping FK(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 11, 0), '早盘   突尼斯杯   比瑟汀(主队) VS 突尼斯施塔德(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 14, 45), '早盘   欧足协国际联赛   Czech Republic(主队) VS 瑞士(客队)', '上半场 - 独赢', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0], ['足球', datetime.datetime(2022, 6, 18, 6, 0), '早盘   日本J联赛   Kawasaki Frontale(主队) VS Hokkaido Consadole Sapporo(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 10, 14, 45), '早盘   爱尔兰甲级联赛   韦克斯福德(主队) VS Treaty United(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 11, 0), '早盘   国际青年U23 AFC Asian Cup   泰国(主队) VS 越南(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 11, 23, 15, 0), '早盘   国际世界杯   比利时(主队) VS 加拿大(客队)', '上半场 - 大/小', 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 11, 28, 12, 0), '早盘   国际世界杯   巴西(主队) VS 瑞士(客队)', '上半场 - 让球', 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 8, 11, 30), '早盘   捷克甲级联赛   齐德利纳河畔赫卢梅茨(主队) VS 布拉格波希米扬斯1905 B队(客队)', '上半场 - 大/小', 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 12, 0), '早盘   欧足协国际联赛   塞浦路斯(主队) VS 科索沃(客队)', '上半场 - 大/小', 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 19, 30), '早盘   美国美足联乙级联赛   南佐治亚州风暴二队(主队) VS 桃树市MOBA(客队)', '大/小', 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 14, 30), '早盘   国际青年U21欧洲锦标赛预选赛   圣马力诺(主队) VS 波兰(客队)', '上半场 - 大/小', 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 7, 9, 15, 0), '早盘   国际欧洲锦标赛，女子   荷兰(主队) VS 瑞典(客队)', '让球', 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 10, 5, 0), '早盘   澳大利亚全国超级联赛,新南威尔士   曼立联(主队) VS 悉尼青年(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 7, 12, 0), '早盘   欧足协国际联赛   芬兰(主队) VS 黑山(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 2, 11, 0), '早盘   国际青年U21欧洲锦标赛预选赛   北马其顿(主队) VS 亚美尼亚(客队)', '上半场 - 让球', 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 29, 14, 0), '早盘   国际俱乐部解放者杯   沙士菲(主队) VS 河床(客队)', '上半场 - 大/小', 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 29, 14, 0), '早盘   国际俱乐部解放者杯   CS Emelec(主队) VS Atletico Mineiro MG(客队)', '大/小', 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 11, 24, 6, 0), '早盘   国际世界杯   瑞士(主队) VS 喀麦隆(客队)', '大/小', 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 13, 2, 0), '早盘   澳大利亚全国超级联赛,新南威尔士   悉尼奥林匹克(主队) VS 卧龙岗狼队(客队)', '独赢', 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0], ['足球', datetime.datetime(2022, 6, 27, 13, 0), '早盘   瑞典超级联赛   Malmo FF(主队) VS 赫尔辛堡(客队)', '上半场 - 大/小', 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0]]
#
#

# content_A1=worksheet.cell(1,1).value
# if content_A1==None:
#     for i in range(1,len(ball_list)):
#         worksheet.cell(1,i,ball_list[i-1])
#     for j in range(2,len(list)):
#         if worksheet.cell(j,1).value==None:
#             worksheet.cell(j, 1, j - 1)
#             for k in range(2, len(ball_list)):
#                 if worksheet.cell(j, k).value == None:
#                     # print(list[k - 2][:]
#                     tt = worksheet.cell(j, k, list[j - 2][k - 2])
#                 else:
#                     pass
#     print("写入完成")
# else:
#     rows = worksheet.max_row
#     columns = worksheet.max_column
#     for i in range(1,int(rows)+1):
#         for j in range(1,int(columns)+1):
#             worksheet.cell(i, j).value=""
#     print("已有内容无法写入,已清空")

# workbook.save(filename="C:\\test\\test_ball.xlsx")
# workbook.close()

list_yy=['1.990', '1.830', '3.600', '3.350', '6.350', '13.950']
AB_list=["A","B","C","D","E","F",]
AB_dict={'A':[],'B':[],'C':[],'D':[],'E':[],'F':[]}

for i in range(0, len(list_yy)):
    AB_dict[AB_list[i]] = list_yy[i]
print(AB_dict)










# workbook.close()
# workbook.save(filename='C:\\test\\test_ball.xlsx')















