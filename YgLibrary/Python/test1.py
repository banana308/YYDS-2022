import pytest
import allure
import openpyxl,xlrd
from os.path import  join,abspath,dirname
import requests,json,time
import openpyxl

#D:\YYDS-2022\YgLibrary\Python
#pytest test1.py -s -q --alluredir=./result --clean-alluredir
#allure serve ./result
#C:\python\Scripts

workbook = openpyxl.load_workbook(r"C:\test\test.xlsx")
#print("读取表格成功")
shenames = workbook.get_sheet_names()
shenames=workbook.sheetnames
print("整个表格所有表名:",shenames)  #['各省市', '测试表']

#还可以通过如下写法获得表对象
worksheet=workbook[shenames[0]]
print("获取所在表的对象：",worksheet)  #<Worksheet "测试表">

worksheet01=workbook[shenames[1]]
print("获取所在表的对象：",worksheet01)  #<Worksheet "测试表">

# name = worksheet.title  # 获取表名
# print("获取所在表的表名:",name)  # 各省市
# 获取该表相应的行数和列数
rows = worksheet.max_row
columns = worksheet.max_column
print("所在表组成有："+str(rows)+"行",str(columns)+"列")  # 32 13

# 获取该表相应的行数和列数
rows01 = worksheet01.max_row
columns01 = worksheet01.max_column
print("所在表组成有："+str(rows01)+"行",str(columns01)+"列")  # 32 13


# url=worksheet["A2"].value
# print(url)
URL=worksheet.cell(row=2,column=1).value
print(URL)

def test_01():
    allure.dynamic.title(worksheet01.cell(row=2,column=2).value)
    with allure.step(worksheet01.cell(row=2,column=3).value):
        url=URL+(worksheet01.cell(row=2,column=5).value)
        header = {'content-type': 'application/json'}
        para=worksheet01.cell(row=2,column=4).value
        # print("请求方式:", para)
        data=json.loads(worksheet01.cell(row=2,column=6).value)
        # print("请求header参数：",data)
        if para=="POST":
            response=requests.post(url=url, headers=header, json=data)
        else:
            response=requests.get(url=url,headers=header,params=data)
        #响应的json数据转换为可被python识别的数据类型
        results = json.loads(response.text)
        # print(results)
        code=results['data']['code']
        worksheet01.cell(row=2, column=7).value = str(results)
        workbook.save
        assert code == 0

def test_02():
    allure.dynamic.title(worksheet01.cell(row=3, column=2).value)
    with allure.step(worksheet01.cell(row=3, column=3).value):
        url = URL + (worksheet01.cell(row=2, column=5).value)
        header = {'content-type': 'application/json'}
        para = worksheet01.cell(row=2, column=4).value
        # print("请求方式:", para)
        data = json.loads(worksheet01.cell(row=2, column=6).value)
        # print("请求header参数：",data)
        if para == "POST":
            response = requests.post(url=url, headers=header, json=data)
        else:
            response = requests.get(url=url, headers=header, params=data)
        # 响应的json数据转换为可被python识别的数据类型
        results = json.loads(response.text)
        # print(results)
        code = results['data']['code']
        worksheet01.cell(row=3, column=7).value = str(results)
        workbook.save
        assert code == 0










